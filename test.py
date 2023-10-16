import tkinter as tk
import openpyxl

# Function to handle the button click event
def calculate_and_save():
    user_values = [entry_values[i].get() for i in range(len(strings))]
    
    # Check if all input fields have values
    if all(user_values):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write headers to Excel
        sheet["A1"] = "String"
        sheet["B1"] = "User Input"

        # Write data to Excel
        for i, s in enumerate(strings):
            sheet.cell(row=i + 2, column=1, value=s)
            sheet.cell(row=i + 2, column=2, value=user_values[i])

        # Save the Excel file
        workbook.save("user_data.xlsx")

        # Close the window after saving the data
        window.quit()
    else:
        result_label.config(text='Please provide a value for each string.')

# Create the main window
window = tk.Tk()
window.title("User Input for Strings")

# Set window size
window.geometry("400x300")

# List of strings
strings = ["String 1", "String 2", "String 3", "String 4"]

# Create labels and entry fields for each string
entry_values = []
for s in strings:
    label = tk.Label(window, text=s)
    label.pack(side="left")
    entry = tk.Entry(window)
    entry.pack(side="left")
    entry_values.append(entry)
    label = tk.Label(window, text="")  # Add an empty label for spacing
    label.pack(side="left")

# Create a save button
save_button = tk.Button(window, text="Save and Close", command=calculate_and_save)
save_button.pack()

# Create a label to display an error message if needed
result_label = tk.Label(window, text="")
result_label.pack()

# Start the main event loop
window.mainloop()
