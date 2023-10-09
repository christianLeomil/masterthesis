import pandas as pd
import classes
import sys
import pandas as pd
from classes import control
from openpyxl import Workbook, utils
from openpyxl.styles import PatternFill, Font, NamedStyle, Alignment
import warnings
import matplotlib.pyplot as plt
import numpy as np
import os


def aux_creator(df_elements):
    list_elements = []
    list_type = []
    for i in df_elements.index:
        for j in range(0,df_elements['# components'].loc[i]):
            list_type.append(i) #entre esses parenteses estava soh i, eu mudei pra class type. pra resolver o problema, tem que mudar substituicao. Solucao mais facil talvez seja renomear parametros da bateria 
            name_element = i + str(j+1)
            list_elements.append(name_element)

    #checking if there is any demand in the inputs, otherwise, print error message
    if not any('demand' in s for s in list_elements):
        print('\n==========ERROR==========')
        print('Model must have a demand in order to run optimization\n')
        sys.exit()

    df_aux = pd.DataFrame({'element': list_elements,
                           'type':list_type})
    
    list_electric_load = []
    list_electric_source = []
    list_thermal_load = []
    list_thermal_source = []    

    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        
        myClass = getattr(classes,element_type)
        component_type = getattr(myClass,'component_type')

        if component_type['electric_load'] == 'yes':
            list_electric_load.append(element)
        if component_type['electric_source'] == 'yes':
            list_electric_source.append(element)
        if component_type['thermal_load'] == 'yes':
            list_thermal_load.append(element)
        if component_type['thermal_source'] == 'yes':
            list_thermal_source.append(element)

    df_con_electric = pd.DataFrame(0,columns = list_electric_source, index = list_electric_load)
    df_con_thermal = pd.DataFrame(0,columns = list_thermal_source, index = list_thermal_load)

    return  df_con_electric, df_con_thermal, df_aux
        
def write_excel(df,path, name_sheet, name_file, boolean):
    with pd.ExcelWriter(path + name_file,mode = 'a',engine = 'openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer,sheet_name = name_sheet, index = boolean)

def connection_creator(df_con_electric, df_con_thermal):

    list_attr_classes = []
    list_attr_classes = list_attr_classes + df_con_electric.index.to_list()
    list_attr_classes = list_attr_classes + df_con_electric.columns.to_list()

    list_attr_classes = list_attr_classes + df_con_thermal.index.to_list()
    list_attr_classes = list_attr_classes + df_con_thermal.columns.to_list()

    #building variable names inside connection matrix ELECTRIC
    for i in df_con_electric.index: 
        for j in df_con_electric.columns:
            if df_con_electric.loc[i,j] != 0:
                if i == j:
                    print('\n==========ERROR==========')
                    print('Energy cannot be transferred from ' + i + ' to ' + j +'\n')
                    sys.exit()
                else:
                    df_con_electric.loc[i,j] = 'P_' + j + '_' + i

    #building variable names inside connection matrix THERMAL
    for i in df_con_thermal.index: 
        for j in df_con_thermal.columns:
            if df_con_thermal.loc[i,j] != 0:
                if i == j:
                    print('\n==========ERROR==========')
                    print('Energy cannot be transferred from ' + i + ' to ' + j +'\n')
                    sys.exit()
                else:
                    df_con_thermal.loc[i,j] = 'Q_' + j + '_' + i

    #creating variable names derived from columns of conection matrix
    list_sub = ['P_from_' + i for i in df_con_electric.columns]
    df_con_electric.columns = list_sub
    list_sub = ['Q_from_' + i for i in df_con_thermal.columns]
    df_con_thermal.columns = list_sub

    #creating variable names derived from index of conection matrix
    list_sub = ['P_to_' + i for i in df_con_electric.index]
    list_sub = [s.replace('P_to_demand','param_P_to_demand') for s in list_sub]
    list_sub = [s.replace('P_to_charging_station','param_P_to_charging_station') for s in list_sub]

    # print('\n->->->->->-list_sub_electric')
    # print(list_sub)

    df_con_electric.index = list_sub

    list_sub = ['Q_to_' + i for i in df_con_thermal.index]
    list_sub = [s.replace('Q_to_demand','param_Q_to_demand') for s in list_sub]

    # print('\n->->->->->-list_sub_thermal')
    # print(list_sub)

    df_con_thermal.index = list_sub


    # creating equations in the 'to' direction ELECTRIC
    list_expressions = []
    for i in df_con_electric.index:
        list_exp_partial = 'model.' + i + '[t] == 0'
        for j in df_con_electric.columns:
            if df_con_electric.loc[i,j] != 0:
                list_exp_partial = list_exp_partial + ' + model.' + df_con_electric.loc[i,j] + '[t]'
        list_expressions.append(list_exp_partial)

    # creating equations in the 'from' direction ELECTRIC
    for i in df_con_electric.columns:
        list_exp_partial = 'model.' + i + '[t] == 0'
        for j in df_con_electric.index:
            if df_con_electric.loc[j,i] != 0:
                list_exp_partial = list_exp_partial + ' + model.' + df_con_electric.loc[j,i] + '[t]'
        list_expressions.append(list_exp_partial)

    # creating equation in the 'to' direction THERMAL
    for i in df_con_thermal.index:
        list_exp_partial = 'model.' + i + '[t] == 0'
        for j in df_con_thermal.columns:
            if df_con_thermal.loc[i,j] != 0:
                list_exp_partial = list_exp_partial + ' + model.' + df_con_thermal.loc[i,j] + '[t]'
        list_expressions.append(list_exp_partial)

    # creating equation in the 'from' direction THERMAL
    for i in df_con_thermal.columns:
        list_exp_partial = 'model.' + i + '[t] == 0'
        for j in df_con_thermal.index:
            if df_con_thermal.loc[j,i] != 0:
                list_exp_partial = list_exp_partial + ' + model.' + df_con_thermal.loc[j,i] + '[t]'
        list_expressions.append(list_exp_partial)

    # creating list with variables that need to be created in the abstract model ELECTRIC
    list_con_variables = []
    list_con_variables = list_con_variables + df_con_electric.columns.to_list()
    list_con_variables = list_con_variables + df_con_electric.index.to_list()
    for i in df_con_electric.index:
        for j in df_con_electric.columns:
            if df_con_electric.loc[i,j] != 0:
                list_con_variables.append(df_con_electric.loc[i,j])
    list_con_variables = list(set(list_con_variables))

    # creating list with variables that need to be created in the abstract model THERMAL
    list_con_variables = list_con_variables + df_con_thermal.columns.to_list()
    list_con_variables = list_con_variables + df_con_thermal.index.to_list()
    for i in df_con_thermal.index:
        for j in df_con_thermal.columns:
            if df_con_thermal.loc[i,j] != 0:
                list_con_variables.append(df_con_thermal.loc[i,j])
    list_con_variables = list(set(list_con_variables))

    return df_con_thermal, df_con_electric, list_expressions, list_con_variables, list_attr_classes


def revenue_constraint_creator(df_con_electric, df_con_thermal):

    df_con_electric = df_con_electric.filter(like = 'P_to_net', axis = 0)
    df_con_electric.columns = [s.replace('P_from_','') for s in df_con_electric.columns]
    df_con_electric.index = [s.replace('P_to_','') for s in df_con_electric.index]


    df_con_thermal = df_con_thermal.filter(like = 'Q_to_net', axis = 0)
    df_con_thermal.columns = [s.replace('Q_from_','') for s in df_con_thermal.columns]
    df_con_thermal.index = [s.replace('Q_to_','') for s in df_con_thermal.index]

    #looping through the connection matrices and checking which elements are connected to the grid
    list_expressions = []
    list_variables = []

    #looping through electric matrix
    for i in df_con_electric.index:
        for j in df_con_electric.columns:
            if df_con_electric.loc[i,j] != 0:
                energy_flux = df_con_electric.loc[i,j]
                #appending expressions and variables related to compensation values
                list_expressions.append(f"model.P_comp_{j}[t] == model.{energy_flux}[t] * model.param_{j}_compensation",)
                list_variables.append(f"P_comp_{j}")
                #appending expressions and variables related to revenue values
                list_expressions.append(f"model.P_rev_{j}[t] == model.{energy_flux}[t] * model.param_{i}_cost_sell_electric")
                list_variables.append(f"P_rev_{j}")

    #looping through thermal matrix
    for i in df_con_thermal.index:
        for j in df_con_thermal.columns:
            if df_con_thermal.loc[i,j] != 0:
                energy_flux = df_con_thermal.loc[i,j]
                #appending expressions and variables related to compensation values
                list_expressions.append(f"model.Q_comp_{j}[t] == model.{energy_flux}[t] * model.param_{j}_compensation",)
                list_variables.append(f"Q_comp_{j}")
                #appending expressions and variables related to revenue values
                list_expressions.append(f"model.Q_rev_{j}[t] == model.{energy_flux}[t] * model.param_{i}_cost_sell_thermal")
                list_variables.append(f"Q_rev_{j}")

    df_expressions_revenue = {'expressions':list_expressions,
                              'variables':list_variables}

    return df_expressions_revenue


def objective_expression_creator(df_aux, df_expressions_revenue):
    #total revenue expressions
    #starting to see if any class has its own 
    list_revenue_total = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        method = getattr(classes,element_type)
        if hasattr(method, 'constraint_revenue'):
            if list_revenue_total == []:
                list_revenue_total.append('model.total_revenue[t] == model.' + element + '_revenue[t]')
            else:
                list_revenue_total[-1] = list_revenue_total[-1] + ' + model.' + element + '_revenue[t]'
    
    print(f"List revenue total looks like this: {list_revenue_total}")
    #now adding revenue and compensation variables created in "revenue_constraint_creator"
    for i in df_expressions_revenue['variables']:
        if list_revenue_total == []:
            list_revenue_total.append('model.total_revenue[t] == model.' + i + '[t]')
        else:
            list_revenue_total[-1] = list_revenue_total[-1] + ' + model.' + i + '[t]'

    print('\n======================================This is the total revenue equation')
    print(list_revenue_total)

    #total investment cost expression
    list_investment_costs_total = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        method = getattr(classes,element_type)
        if hasattr(method,'constraint_investment_costs'):
            if list_investment_costs_total == []:
                list_investment_costs_total.append('model.total_investment_costs[t] == ' + ' model.' + element + '_inv_cost[t]')
            else:
                list_investment_costs_total[-1] = list_investment_costs_total[-1] + ' + model.'+ element + '_inv_cost[t]'

    print('\n======================================This is the total investment costs')
    print(list_investment_costs_total)


    #total operation costs expression
    list_operation_costs_total = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        method = getattr(classes,element_type)
        if hasattr(method,'constraint_operation_costs'):
            if list_operation_costs_total == [] :
                list_operation_costs_total.append('model.total_operation_costs[t] == ' + ' model.' + element + '_op_cost[t]')
            else:
                list_operation_costs_total[-1] = list_operation_costs_total[-1] + ' + model.'+ element + '_op_cost[t]'

    print('\n======================================This is the total operation costs')
    print(list_operation_costs_total)


    #total emissions expressions
    list_emissions_total = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        method = getattr(classes,element_type)
        if hasattr(method,'constraint_emissions'):
            if list_emissions_total == [] :
                list_emissions_total.append('model.total_emissions[t] == ' + 'model.' + element +'_emissions[t]')
            else:
                list_emissions_total[-1] = list_emissions_total[-1] + ' + model.' + element +'_emissions[t]'

    print('\n======================================This is the total emission costs')
    print(list_emissions_total)

    return list_revenue_total, list_operation_costs_total, list_investment_costs_total, list_emissions_total

def objective_constraint_creator(df_aux): # this function creates the constraints for the objective function to work

    # creating constraints to calculate bouhgt and sold energy from the grid.
    list_buy_constraint = ['model.total_buy[t] == '] 
    list_sell_constraint = ['model.total_sell[t] == ']

    df_temp = df_aux[df_aux['type'] == 'net' ].reset_index(drop = True)
    
    for i in df_temp.index:
        element = df_temp['element'].iloc[i]
        if i == 0:
            list_buy_constraint[-1] = list_buy_constraint[-1] + ' model.' + element + '_buy_electric[t]' + ' + model.' + element + '_buy_thermal[t]'
            list_sell_constraint[-1] = list_sell_constraint[-1] + ' model.' + element + '_sell_electric[t]' + ' + model.' + element + '_sell_thermal[t]'
        else:
            list_buy_constraint[-1] = list_buy_constraint[-1] + ' + model.' + element + '_buy_electric[t]' + ' + model.' + element + '_buy_thermal[t]'
            list_sell_constraint[-1] = list_sell_constraint[-1] + ' + model.' + element + '_sell_electric[t]' + ' + model.' + element + '_sell_thermal[t]'
    
    list_objective_constraints = list_buy_constraint + list_sell_constraint


    list_operation_costs_total = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        method = getattr(classes,element_type)
        if hasattr(method,'constraint_operation_costs'):
            if list_operation_costs_total == [] :
                list_operation_costs_total.append('model.total_operation_cost[t] == ' + ' model.' + element + '_op_cost[t]')
            else:
                list_operation_costs_total[-1] = list_operation_costs_total[-1] + '+ model.'+ element + '_op_cost[t]'

    list_objective_constraints = list_objective_constraints + list_operation_costs_total

    list_emissions_constraint = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        method = getattr(classes,element_type)
        if hasattr(method,'constraint_emissions'):
            if list_emissions_constraint == [] :
                list_emissions_constraint.append('model.total_emissions[t] == ' + 'model. ' + element +'_emissions[t]')
            else:
                list_emissions_constraint[-1] = list_emissions_constraint[-1] + ' + model. ' + element +'_emissions[t]'
    
    # print('\nEmission Constriants') 
    # print(list_emissions_constraint)

    list_objective_constraints = list_objective_constraints + list_emissions_constraint

    return list_objective_constraints

def organize_output_columns(df_variable_values,df_aux):
    list_columns = df_variable_values.columns
    list_temp = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        for j in list_columns:
            if '_'+ element +'_' in j:
                list_temp.append(j)
            elif element +'_' in j:
                list_temp.append(j)
            elif 'from_'+ element in j:
                list_temp.append(j)
            elif 'to_' + element in j:
                list_temp.append(j)

    complimentary_list = list(set(list_columns) - set(list_temp))
    list_temp = list_temp + complimentary_list
    list_temp.remove('TimeStep')
    list_temp.insert(0,'TimeStep')
    df_variable_values = df_variable_values.reindex(columns = list_temp)

    return df_variable_values

def breaking_dataframe(df,horizon,saved_position):
    list_split = []
    for i in range(0,len(df),saved_position):
        # print('\n------------'+str(i))
        if i == 0:
            list_split.append(df.iloc[i : i + horizon])
        else:
            list_split.append(df.iloc[i-1 : i+horizon])
        # print(list_split[-1])
    return list_split

def save_variables_last_time_step(df_input_others,df_variables_last_time_step):
    for j in df_variables_last_time_step.index:
        if df_variables_last_time_step['Parameter'].iloc[j] in df_input_others['Parameter'].tolist():
            df_input_others.loc[df_input_others['Parameter']==df_variables_last_time_step['Parameter'].iloc[j], 'Value'] = df_variables_last_time_step['Value'].iloc[j]
        else:
            row_to_append = df_variables_last_time_step.loc[j]
            df_input_others = df_input_others.append(row_to_append, ignore_index = True)
            
    return df_input_others

def financial_analysis(control):

    warnings.filterwarnings("ignore")

    path_output = control.path_output

    df_final = pd.read_excel(path_output + 'df_final.xlsx')
    reference_date = control.reference_date

    df_final['Date'] = reference_date + pd.to_timedelta(df_final['TimeStep'], unit = 'h')
    df_final.set_index('Date', inplace = True)

    df_final['Year'] = df_final.index.year.astype(int)
    df_final['Week'] = df_final.index.week.astype(int)
    df_final['Year Month'] = df_final.index.strftime('%Y-%m')
    df_final['Year Week'] = df_final.index.strftime('%Y-%W').astype(str)
    df_final['Year Week'] = [ s.replace('-', ' - wk ') for s in df_final['Year Week']]

    weekly_sum_df = df_final.groupby('Year Week').sum()
    monthly_sum_df = df_final.groupby('Year Month').sum()
    annual_sum_df = df_final.groupby('Year').sum()

    # Create a new Excel workbook
    wb = Workbook()

    # Select the active sheet (you can also create a new one if needed)
    sheet = wb.active

    #some formats
    thousands_separator_format = NamedStyle(name="thousands_separator_format")
    thousands_separator_format.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'

    centered_style = NamedStyle(name="centered_style")
    centered_style.alignment = Alignment(horizontal = "center", vertical = "center")

    normal_style = NamedStyle(name="normal_style")
    normal_style.alignment = Alignment(horizontal = "left", vertical = "center", indent = 0)

    normal_style_with_indent = NamedStyle(name="normal_style_with_indent")
    normal_style_with_indent.alignment = Alignment(horizontal = "left", vertical = "center", indent = 2)


    # function for formatting cells
    def formatting_cells(minRow, maxRow, minCol, maxCol, colorCode, boldOption, fontOption, sizeOption, styleOption):
        for row in sheet.iter_rows(min_row = minRow, max_row = maxRow, min_col = minCol, max_col = maxCol):
            for cell in row:
                cell.style = styleOption
                cell.fill = PatternFill(start_color = colorCode, end_color=colorCode, fill_type="solid")
                cell.font = Font(bold = boldOption, size = sizeOption, name = fontOption)
        return


    def conditional_formatting_cells(minRow, maxRow, minCol, maxCol, colorCode, boldOption, fontOption, sizeOption, styleOption,fontColor):
        for cell in sheet.iter_rows(min_row = minRow, max_row = maxRow, min_col = minCol, max_col = maxCol):
            for c in cell:
                c.style = styleOption
                c.fill = PatternFill(start_color = colorCode, end_color = colorCode, fill_type="solid")
                if c.value >= 0:
                    c.font = Font(bold = boldOption, size = sizeOption, name = fontOption)
                else:
                    c.font = Font(bold = boldOption, size = sizeOption, name = fontOption, color = fontColor)
        return

    def set_column_width(width, index):
        column_index = index
        column_letter = utils.get_column_letter(column_index)
        column_dimension = sheet.column_dimensions[column_letter]
        column_dimension.width = width
        return

    # function for inserting values

    # Inserting values and formating cells from the top of the sheet to down
    # --------------------------------------------------------------------------
    # region top strips

    sheet.cell(row = 2, column = 2, value = 'Financial Analysis - Calculations (EoP)')
    sheet.cell(row = 3, column = 2, value = 'Units: €')

    formatting_cells(minRow = 2, maxRow = 2, minCol = 1, maxCol = 1000, colorCode = "009EE3", 
                    boldOption = True, sizeOption = 11, fontOption = 'Arial', styleOption= normal_style)

    formatting_cells(minRow = 3, maxRow = 3, minCol = 1, maxCol = 1000, colorCode = "93D3F7", 
                    boldOption = False, sizeOption = 11, fontOption = 'Arial', styleOption= normal_style)

    # endregion
    # --------------------------------------------------------------------------
    # region inserting operation costs

    col_index_yearly_data = 5
    col_index_monthly_data = 5 + len(annual_sum_df) + 1
    col_index_weekly_data = 5 + len(annual_sum_df) + 1 + len(monthly_sum_df) + 1

    # operation costs title
    sheet.cell(row = 7, column = 3, value = '1 - Operation Costs')

    #formating row operation costs
    formatting_cells(minRow = 7, maxRow = 7, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)

    # Year, month and weeks title
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_yearly_data, maxCol = col_index_yearly_data + len(annual_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_monthly_data, maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_weekly_data, maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)


    # inserting and formatting ANNUAL DATA
    list_columns = [s for s in annual_sum_df if 'op_cost' in s or "buy" in s or "sell" in s]
    annual_costs_df = annual_sum_df[list_columns]
    annual_costs_df *= -1 

    #inserting title of row columns
    for idx, value in enumerate(annual_costs_df.columns, start = 8):
        sheet.cell(row = idx, column = 3, value = value)

    #inserting years titles
    for idx, value in enumerate(annual_costs_df.index, start = 5):
        sheet.cell(row = 5, column = idx, value = value)

    #inserting values of yearly data
    for row_idx, row in enumerate(annual_costs_df.iterrows(), start = 5):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    #formatting titles of row columns
    formatting_cells(minRow = 8, maxRow =  8 + len(annual_costs_df.columns) - 1, minCol = 1, maxCol = 3, colorCode = "D4EEFC", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption= normal_style_with_indent)

    #formatting numbers of yearly values
    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(annual_costs_df.columns) - 1, minCol = 5, maxCol = 5 + len(annual_sum_df) - 1, 
                                    colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")

    # formatting annual sum for the years
    list_annual_sum = annual_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_annual_sum, start = 5):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = 5, maxCol = 5 + len(list_annual_sum) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")


    # inserting and formatting MONTHLY DATA
    list_columns = [s for s in monthly_sum_df if 'op_cost' in s or "buy" in s or "sell" in s]
    monthly_costs_df = monthly_sum_df[list_columns]
    monthly_costs_df *= -1 

    #inserting month titles
    for idx, value in enumerate(monthly_costs_df.index, start = col_index_monthly_data):
        sheet.cell(row = 5, column = idx, value = value)

    for row_idx, row in enumerate(monthly_costs_df.iterrows(), start = col_index_monthly_data):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(annual_costs_df.columns) - 1, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # formatting annual sum for the months
    list_monthly_sum = monthly_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_monthly_sum, start = col_index_monthly_data):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting monthly sum for the months
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")


    # inserting and formatting WEEKLY DATA
    list_columns = [s for s in weekly_sum_df if 'op_cost' in s or "buy" in s or "sell" in s]
    weekly_costs_df = weekly_sum_df[list_columns]
    weekly_costs_df *= -1 

    #inserting week titles
    for idx, value in enumerate(weekly_costs_df.index, start = col_index_weekly_data):
        sheet.cell(row = 5, column = idx, value = value)

    for row_idx, row in enumerate(weekly_costs_df.iterrows(), start = col_index_weekly_data):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(annual_costs_df.columns) - 1, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # formatting annual sum for the weeks
    list_weekly_sum = weekly_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_weekly_sum, start = col_index_weekly_data):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting monthly sum for the weeks
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")


    # endregion
    # --------------------------------------------------------------------------
    # region investment costs

    row_index_inv_data = 7 + 2 + len(annual_costs_df.columns)

    # operation costs title
    sheet.cell(row = row_index_inv_data, column = 3, value = '2 - Investment Costs')

    #formating row operation costs
    formatting_cells(minRow = row_index_inv_data, maxRow = row_index_inv_data, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)

    # inserting and formatting ANNUAL DATA
    list_columns = [s for s in annual_sum_df if 'inv_cost' in s]
    annual_costs_df = annual_sum_df[list_columns]
    annual_costs_df *= -1 

    #inserting title of row columns
    for idx, value in enumerate(annual_costs_df.columns, start = row_index_inv_data + 1):
        sheet.cell(row = idx, column = 3, value = value)

    #inserting values of yearly data
    for row_idx, row in enumerate(annual_costs_df.iterrows(), start = 5):
        for c_idx, value in enumerate(row[1], start = row_index_inv_data + 1):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    #formatting titles of row columns
    formatting_cells(minRow = row_index_inv_data + 1, maxRow =  row_index_inv_data + 1 + len(annual_costs_df.columns) - 1, minCol = 1, maxCol = 3, colorCode = "D4EEFC", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style_with_indent)

    #formatting numbers of yearly values
    conditional_formatting_cells(minRow = row_index_inv_data + 1 , maxRow  = row_index_inv_data + 1  + len(annual_costs_df.columns) - 1, minCol = 5, maxCol = 5 + len(annual_sum_df) - 1, 
                                    colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")

    # formatting annual sum for the years
    list_annual_sum_inv = annual_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_annual_sum_inv, start = 5):
        sheet.cell(row = row_index_inv_data, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = row_index_inv_data, maxRow  = row_index_inv_data, minCol = 5, maxCol = 5 + len(list_annual_sum_inv) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")



    # inserting and formatting MONTHLY DATA
    list_columns = [s for s in monthly_sum_df if 'inv_cost' in s]
    monthly_costs_df = monthly_sum_df[list_columns]
    monthly_costs_df *= -1 

    #inserting data for monthly values
    for row_idx, row in enumerate(monthly_costs_df.iterrows(), start = col_index_monthly_data):
        for c_idx, value in enumerate(row[1], start = row_index_inv_data + 1 ):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = row_index_inv_data + 1 , maxRow  =  row_index_inv_data + 1 + len(annual_costs_df.columns) - 1, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # inserting annual sum for the months
    list_monthly_sum_inv = monthly_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_monthly_sum_inv, start = col_index_monthly_data):
        sheet.cell(row =  row_index_inv_data, column = col_idx, value = value)

    # formatting monthly sum for the years
    conditional_formatting_cells(minRow =  row_index_inv_data, maxRow  =  row_index_inv_data, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum_inv) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")


    # inserting and formatting WEEKLY DATA
    list_columns = [s for s in weekly_sum_df if 'inv_cost' in s]
    weekly_costs_df = weekly_sum_df[list_columns]
    weekly_costs_df *= -1 


    for row_idx, row in enumerate(weekly_costs_df.iterrows(), start = col_index_weekly_data):
        for c_idx, value in enumerate(row[1], start = row_index_inv_data + 1):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = row_index_inv_data + 1 , maxRow  = row_index_inv_data + 1 + len(annual_costs_df.columns) - 1, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # inserting total sum for the weeks
    list_weekly_sum_inv = weekly_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_weekly_sum_inv, start = col_index_weekly_data):
        sheet.cell(row = row_index_inv_data, column = col_idx, value = value)

    # formatting total sum for the weeks
    conditional_formatting_cells(minRow = row_index_inv_data, maxRow  = row_index_inv_data, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(list_weekly_sum) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")


    # endregion
    # --------------------------------------------------------------------------
    # region other adjustments

    set_column_width(23, 3)
    set_column_width(2, 4)
    set_column_width(2, col_index_monthly_data - 1)
    set_column_width(2, col_index_weekly_data -1)

    for i in range(col_index_weekly_data, col_index_weekly_data + len(weekly_costs_df)):
        set_column_width(13, i)

    # endregion
    # --------------------------------------------------------------------------
    # region other adjustments

    row_index_total = row_index_inv_data + len(annual_costs_df.columns) + 1 + 1

    # operation costs title
    sheet.cell(row = row_index_total, column = 3, value = '3 - Result')

    #formating row operation costs
    formatting_cells(minRow = row_index_total, maxRow = row_index_total, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)

    annual_result  = np.array(list_annual_sum) + np.array(list_annual_sum_inv)
    monthly_result  = np.array(list_monthly_sum) + np.array(list_monthly_sum_inv)
    weekly_result  = np.array(list_weekly_sum) + np.array(list_weekly_sum_inv)


    # inserting annual result for the years
    for col_idx, value in enumerate(annual_result, start = 5):
        sheet.cell(row = row_index_total, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = row_index_total, maxRow  = row_index_total, minCol = 5, maxCol = 5 + len(list_annual_sum_inv) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")


    # inserting annual sum for the months
    for col_idx, value in enumerate(monthly_result, start = col_index_monthly_data):
        sheet.cell(row =  row_index_total, column = col_idx, value = value)

    # formatting monthly sum for the years
    conditional_formatting_cells(minRow =  row_index_total, maxRow  =  row_index_total, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum_inv) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")


    # inserting total sum for the weeks
    for col_idx, value in enumerate(weekly_result, start = col_index_weekly_data):
        sheet.cell(row = row_index_total, column = col_idx, value = value)

    # formatting total sum for the weeks
    conditional_formatting_cells(minRow = row_index_total, maxRow  = row_index_total, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(list_weekly_sum) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")
    



    # now inserting the cumulated sum of the cash flows
    row_index_total_cumulated = row_index_total + 1 + 1

    # operation costs title
    sheet.cell(row = row_index_total_cumulated, column = 3, value = '4 - Result accumulated')

    #formating row operation costs
    formatting_cells(minRow = row_index_total_cumulated, maxRow = row_index_total_cumulated, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)
    
    
    # inserting annual result for the years
    for col_idx, value in enumerate(np.cumsum(annual_result), start = 5):
        sheet.cell(row = row_index_total_cumulated, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = row_index_total_cumulated, maxRow  = row_index_total_cumulated, minCol = 5, maxCol = 5 + len(list_annual_sum_inv) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")

    # inserting annual sum for the months
    for col_idx, value in enumerate(np.cumsum(monthly_result), start = col_index_monthly_data):
        sheet.cell(row =  row_index_total_cumulated, column = col_idx, value = value)

    # formatting monthly sum for the years
    conditional_formatting_cells(minRow =  row_index_total_cumulated, maxRow  =  row_index_total_cumulated, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum_inv) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # inserting total sum for the weeks
    for col_idx, value in enumerate(np.cumsum(weekly_result), start = col_index_weekly_data):
        sheet.cell(row = row_index_total_cumulated, column = col_idx, value = value)

    # formatting total sum for the weeks
    conditional_formatting_cells(minRow = row_index_total_cumulated, maxRow  = row_index_total_cumulated, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(list_weekly_sum) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    #endregion

    # Hide gridlines
    sheet.sheet_view.showGridLines = False
    mycell = sheet['D6']
    sheet.freeze_panes = mycell

    # Save the Excel file
    wb.save(path_output + "financial_analysis.xlsx")

def emissions_analysis(control):

    warnings.filterwarnings("ignore")

    path_output = control.path_output

    df_final = pd.read_excel(path_output + 'df_final.xlsx')
    reference_date = control.reference_date

    df_final['Date'] = reference_date + pd.to_timedelta(df_final['TimeStep'], unit = 'h')
    df_final.set_index('Date', inplace = True)

    df_final['Year'] = df_final.index.year.astype(int)
    df_final['Week'] = df_final.index.week.astype(int)
    df_final['Year Month'] = df_final.index.strftime('%Y-%m')
    df_final['Year Week'] = df_final.index.strftime('%Y-%W').astype(str)
    df_final['Year Week'] = [s.replace('-', ' - wk ') for s in df_final['Year Week']]

    weekly_sum_df = df_final.groupby('Year Week').sum()
    monthly_sum_df = df_final.groupby('Year Month').sum()
    annual_sum_df = df_final.groupby('Year').sum()

    # Create a new Excel workbook
    wb = Workbook()

    # Select the active sheet (you can also create a new one if needed)
    sheet = wb.active

    #some formats
    thousands_separator_format = NamedStyle(name="thousands_separator_format")
    thousands_separator_format.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'

    centered_style = NamedStyle(name="centered_style")
    centered_style.alignment = Alignment(horizontal = "center", vertical = "center")

    normal_style = NamedStyle(name="normal_style")
    normal_style.alignment = Alignment(horizontal = "left", vertical = "center", indent = 0)

    normal_style_with_indent = NamedStyle(name="normal_style_with_indent")
    normal_style_with_indent.alignment = Alignment(horizontal = "left", vertical = "center", indent = 2)


    # function for formatting cells
    def formatting_cells(minRow, maxRow, minCol, maxCol, colorCode, boldOption, fontOption, sizeOption, styleOption):
        for row in sheet.iter_rows(min_row = minRow, max_row = maxRow, min_col = minCol, max_col = maxCol):
            for cell in row:
                cell.style = styleOption
                cell.fill = PatternFill(start_color = colorCode, end_color=colorCode, fill_type="solid")
                cell.font = Font(bold = boldOption, size = sizeOption, name = fontOption)
        return


    def conditional_formatting_cells(minRow, maxRow, minCol, maxCol, colorCode, boldOption, fontOption, sizeOption, styleOption,fontColor):
        for cell in sheet.iter_rows(min_row = minRow, max_row = maxRow, min_col = minCol, max_col = maxCol):
            for c in cell:
                c.style = styleOption
                c.fill = PatternFill(start_color = colorCode, end_color = colorCode, fill_type="solid")
                if c.value >= 0:
                    c.font = Font(bold = boldOption, size = sizeOption, name = fontOption)
                else:
                    c.font = Font(bold = boldOption, size = sizeOption, name = fontOption, color = fontColor)
        return

    def set_column_width(width, index):
        column_index = index
        column_letter = utils.get_column_letter(column_index)
        column_dimension = sheet.column_dimensions[column_letter]
        column_dimension.width = width
        return

    # function for inserting values

    # Inserting values and formating cells from the top of the sheet to down
    # --------------------------------------------------------------------------
    # region top strips

    sheet.cell(row = 2, column = 2, value = 'Emissions Analysis - Calculations (EoP)')
    sheet.cell(row = 3, column = 2, value = 'Units: kg CO2')

    formatting_cells(minRow = 2, maxRow = 2, minCol = 1, maxCol = 1000, colorCode = "009EE3", 
                    boldOption = True, sizeOption = 11, fontOption = 'Arial', styleOption= normal_style)

    formatting_cells(minRow = 3, maxRow = 3, minCol = 1, maxCol = 1000, colorCode = "93D3F7", 
                    boldOption = False, sizeOption = 11, fontOption = 'Arial', styleOption= normal_style)

    # endregion
    # --------------------------------------------------------------------------
    # region inserting operation costs

    col_index_yearly_data = 5
    col_index_monthly_data = 5 + len(annual_sum_df) + 1
    col_index_weekly_data = 5 + len(annual_sum_df) + 1 + len(monthly_sum_df) + 1

    # operation costs title
    sheet.cell(row = 7, column = 3, value = '1 - Emissions')

    #formating row operation costs
    formatting_cells(minRow = 7, maxRow = 7, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)

    # Year, month and weeks title
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_yearly_data, maxCol = col_index_yearly_data + len(annual_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_monthly_data, maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_weekly_data, maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)


    # inserting and formatting ANNUAL DATA
    list_columns = [s for s in annual_sum_df if 'emissions' in s and 'total' not in s]
    annual_emissions_df = annual_sum_df[list_columns]
    # annual_emissions_df *= -1

    #inserting title of row columns
    for idx, value in enumerate(annual_emissions_df.columns, start = 8):
        sheet.cell(row = idx, column = 3, value = value)

    #inserting years titles
    for idx, value in enumerate(annual_emissions_df.index, start = 5):
        sheet.cell(row = 5, column = idx, value = value)

    #inserting values of yearly data
    for row_idx, row in enumerate(annual_emissions_df.iterrows(), start = 5):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    #formatting titles of row columns
    formatting_cells(minRow = 8, maxRow =  8 + len(annual_emissions_df.columns) - 1, minCol = 1, maxCol = 3, colorCode = "D4EEFC", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption= normal_style_with_indent)

    #formatting numbers of yearly values
    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(annual_emissions_df.columns) - 1, minCol = 5, maxCol = 5 + len(annual_sum_df) - 1, 
                                    colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "00B050")

    # formatting annual sum for the years
    list_annual_sum = annual_emissions_df.sum(axis =1)
    for col_idx, value in enumerate(list_annual_sum, start = 5):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = 5, maxCol = 5 + len(list_annual_sum) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "00B050")


    # inserting and formatting MONTHLY DATA
    list_columns = [s for s in monthly_sum_df if 'emissions' in s and 'total' not in s]
    monthly_emissions_df = monthly_sum_df[list_columns]
    # monthly_emissions_df *= -1 

    #inserting month titles
    for idx, value in enumerate(monthly_emissions_df.index, start = col_index_monthly_data):
        sheet.cell(row = 5, column = idx, value = value)

    for row_idx, row in enumerate(monthly_emissions_df.iterrows(), start = col_index_monthly_data):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(monthly_emissions_df.columns) - 1, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")

    # formatting annual sum for the months
    list_monthly_sum = monthly_emissions_df.sum(axis =1)
    for col_idx, value in enumerate(list_monthly_sum, start = col_index_monthly_data):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting monthly sum for the months
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")


    # inserting and formatting WEEKLY DATA
    list_columns = [s for s in weekly_sum_df if 'emissions' in s and 'total' not in s]
    weekly_emissions_df = weekly_sum_df[list_columns]
    # weekly_costs_df *= -1

    #inserting week titles
    for idx, value in enumerate(weekly_emissions_df.index, start = col_index_weekly_data):
        sheet.cell(row = 5, column = idx, value = value)

    for row_idx, row in enumerate(weekly_emissions_df.iterrows(), start = col_index_weekly_data):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(weekly_emissions_df.columns) - 1, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")

    # formatting annual sum for the weeks
    list_weekly_sum = weekly_emissions_df.sum(axis =1)
    for col_idx, value in enumerate(list_weekly_sum, start = col_index_weekly_data):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting monthly sum for the weeks
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")


    # endregion
    # --------------------------------------------------------------------------
    # region other adjustments

    set_column_width(23, 3)
    set_column_width(2, 4)
    set_column_width(2, col_index_monthly_data - 1)
    set_column_width(2, col_index_weekly_data -1)

    for i in range(col_index_weekly_data, col_index_weekly_data + len(weekly_emissions_df)):
        set_column_width(13, i)

    # endregion
    # --------------------------------------------------------------------------
    # region other adjustments

    annual_result  = np.array(list_annual_sum) 
    monthly_result  = np.array(list_monthly_sum) 
    weekly_result  = np.array(list_weekly_sum) 


    # now inserting the cumulated sum of the cash flows
    row_index_total_cumulated = 7 + len(annual_emissions_df.columns) + 1 + 1

    # operation costs title
    sheet.cell(row = row_index_total_cumulated, column = 3, value = '2 - Emissions accumulated')

    #formating row operation costs
    formatting_cells(minRow = row_index_total_cumulated, maxRow = row_index_total_cumulated, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)
    
    
    # inserting annual result for the years
    for col_idx, value in enumerate(np.cumsum(annual_result), start = 5):
        sheet.cell(row = row_index_total_cumulated, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = row_index_total_cumulated, maxRow  = row_index_total_cumulated, minCol = 5, maxCol = 5 + len(annual_result) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "00B050")

    # inserting annual sum for the months
    for col_idx, value in enumerate(np.cumsum(monthly_result), start = col_index_monthly_data):
        sheet.cell(row =  row_index_total_cumulated, column = col_idx, value = value)

    # formatting monthly sum for the years
    conditional_formatting_cells(minRow =  row_index_total_cumulated, maxRow  =  row_index_total_cumulated, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(monthly_result) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")

    # inserting total sum for the weeks
    for col_idx, value in enumerate(np.cumsum(weekly_result), start = col_index_weekly_data):
        sheet.cell(row = row_index_total_cumulated, column = col_idx, value = value)

    # formatting total sum for the weeks
    conditional_formatting_cells(minRow = row_index_total_cumulated, maxRow  = row_index_total_cumulated, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_result) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")

    #endregion

    # Hide gridlines
    sheet.sheet_view.showGridLines = False
    mycell = sheet['D6']
    sheet.freeze_panes = mycell

    # Save the Excel file
    wb.save(path_output + "emission_analysis.xlsx")

def charts_generator(control,df_aux):
    path_output = control.path_output
    path_charts = control.path_charts
    reference_date = control.reference_date

    df_final = pd.read_excel(path_output + 'df_final.xlsx')
    df_final['Date'] = reference_date + pd.to_timedelta(df_final['TimeStep'], unit = 'h')
    if control.time_span < 100:
        df_final = df_final[:control.time_span]
    else:
        df_final = df_final[:100]

    x_axis = df_final['TimeStep'].tolist()
    figure_size = (15,8)

    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type =  df_aux['type'].iloc[i]

        list_connections_electric = [s for s in df_final.columns if '_' + element + '_' in s and 'P_' in s]
        list_connections_thermal = [s for s in df_final.columns if '_' + element + '_' in s and 'Q_' in s]

        folder_path = path_charts + '/' + element + '/'

        try:
            os.mkdir(folder_path)
        except FileExistsError:
            print(f"Folder '{element}' already exists at {folder_path}")
        except Exception as e:
            print(f"An error occurred: {str(e)}")

        plt.figure(figsize = figure_size)
        bottom = np.zeros(len(x_axis))
        if len(list_connections_electric) > 0:
            for j in list_connections_electric:
                plt.bar(x_axis, df_final[j],bottom = bottom)
                bottom += df_final[j]
            plt.xlabel('Time [hs]')
            plt.ylabel('Power [kW]')
            plt.title(f"Electric power distribution - {element}")
            plt.legend(list_connections_electric)
            plt.savefig(folder_path + 'P - ' + element + '.png')
            plt.close()

        plt.figure(figsize = figure_size)
        bottom = np.zeros(len(x_axis))
        if len(list_connections_thermal) > 0:
            for j in list_connections_thermal:
                plt.bar(x_axis, df_final[j],bottom = bottom)
                bottom += df_final[j]
            plt.xlabel('Time [hs]')
            plt.ylabel('Power [kW]')
            plt.title(f"Thermal power distribution - {element}")
            plt.legend(list_connections_thermal)
            plt.savefig(folder_path +'Q - ' + element + '.png')
            plt.close()

        if element_type == 'bat':
            plt.figure(figsize = figure_size)
            plt.bar(x_axis, df_final[element + '_energy'])
            plt.xlabel('Time [hs]')
            plt.ylabel('Energy [kWh]')
            plt.title(f"Energy level - {element}")
            plt.legend(element)
            plt.savefig(folder_path + 'E - ' + element +'.png')
            plt.close()

        elif element_type == 'bat_with_aging':
            plt.figure(figsize = figure_size)
            plt.bar(x_axis, df_final[element + '_SOC'])
            plt.xlabel('Time [hs]')
            plt.ylabel('State of charge [-]')
            plt.title(f"State of charge - {element}")
            plt.legend(element)
            plt.savefig(folder_path + 'SoC - ' + element +'.png')
            plt.close()

            plt.figure(figsize = figure_size)
            plt.plot(x_axis, df_final[element + '_SOC_max'])
            plt.xlabel('Time [hs]')
            plt.ylabel('max state of charge [-]')
            plt.title(f"Effect of aging on max state of charge - {element}")
            plt.legend(element)
            plt.savefig(folder_path + 'max SoC - ' + element +'.png')
            plt.close()