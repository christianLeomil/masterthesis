import pandas as pd
import classes
import sys
import pandas as pd
from openpyxl import Workbook, utils
from openpyxl.styles import PatternFill, Font, NamedStyle, Alignment
import warnings
import matplotlib.pyplot as plt
import numpy as np
import os
import inspect




def write_excel(df,path, name_sheet, name_file, boolean):
    with pd.ExcelWriter(path + name_file,mode = 'a',engine = 'openpyxl', if_sheet_exists='replace') as writer:
        df.to_excel(writer,sheet_name = name_sheet, index = boolean)

def write_avaliable_elements_and_domain_names(control):
    class_names = [name for name, obj in inspect.getmembers(classes) if inspect.isclass(obj)]
    list_elements = []
    for i in class_names:
        if hasattr(getattr(classes,i),'domain_type'):
            if len(getattr(classes,i).domain_type['energy_domains']) <= control.number_energy_domains:
                list_elements.append(i)
            
    df_elements = pd.DataFrame({'class_type': list_elements,
                               '# components': [0]*len(list_elements)})
    write_excel(df_elements,control.path_input,'microgrid_components', 'input.xlsx', False)

    list_domains = []
    for i in range(control.number_energy_domains):
        list_domains.append('domain' + str(i + 1))
    df_domain = pd.DataFrame(data = [[0] * len(list_domains)], columns = list_domains)
    df_domain = df_domain.T
    df_domain.columns = ['domain_names']
    df_domain.index.name = None
    write_excel(df_domain,control.path_input, 'energy_domains_names', 'input.xlsx', True)

    return

def create_element_df_and_domain_selection_df(df_elements,df_domains,control):  
    #checking if all domains were named, otherwise, print error message
    if all(isinstance(s,str) for s in df_domains['domain_names'].tolist()) and all(s != ' ' for s in df_domains['domain_names'].tolist()):
        pass
    else:
        print('\n==========ERROR==========')
        print('All domains must be named with a string and cannot be empty values\n')
        sys.exit()

    # creating element df
    list_elements = []
    list_type = []
    for i in df_elements.index:
        for j in range(0,df_elements.loc[i,'# components']):
            list_type.append(i)
            name_element = i + str(j+1)
            list_elements.append(name_element)

    #checking if there is any demand in the inputs, otherwise, print error message
    if not any('demand' in s for s in list_elements):
        print('\n==========ERROR==========')
        print('Model must have a demand in order to run optimization\n')
        sys.exit()

    df_aux = pd.DataFrame({'element': list_elements,
                           'type':list_type})
    
    df_aux.to_excel(control.path_output + 'df_aux.xlsx',index = False)
    
    # creating element df
    list_component_domain = []
    list_component_source_domain = []
    list_component_load_domain = []
    for i in df_aux.index:
        # element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        list_component_domain.append(getattr(classes,element_type).domain_type['energy_domains'])
        list_component_source_domain.append(getattr(classes,element_type).domain_type['source_domains'])
        list_component_load_domain.append(getattr(classes,element_type).domain_type['load_domains'])

    df_domain_selection = pd.DataFrame({'element':list_elements,
                                        'component_domains':list_component_domain,
                                        'component_source_domains':list_component_source_domain,
                                        'component_load_domains':list_component_load_domain,
                                        'optimization_domains':[df_domains['domain_names'].tolist()] * len(list_component_domain)})
    
    biggest_list = max(list_component_domain,key = len)
    for i in range(len(biggest_list)):
        list_action = []
        name_column = 'domain_choice' +str(i + 1)
        for j in range(len(df_domain_selection)):
            if len(list_component_domain[j]) >= (i + 1):
                list_action.append('insert here')
            else:
                list_action.append(0)
        df_domain_selection[name_column] = list_action
    write_excel(df_domain_selection,control.path_input, 'domain_selection', 'input.xlsx', False)

    return df_aux

def create_connection_revenue_and_stock_matrices(df_domain, df_domain_selection,control):
    #checking if given domain names match pre-defined names  
    list_columns = [s for s in df_domain_selection.columns if 'domain_choice' in s]
    for i in list_columns:
        for j in df_domain_selection.index:
            if df_domain_selection.loc[j,i] != 0:
                if df_domain_selection.loc[j,i] not in df_domain['domain_names'].tolist():
                    print('\n==========ERROR==========')
                    print(f"'{df_domain_selection.loc[j,i]}' is not one of the defined domain names. Please start again.")
                    sys.exit()

    #checking if element has the same domains inserted for different energy types.
    df_temp = df_domain_selection[list_columns].copy()
    for i in range(len(df_temp)):
        if df_temp.iloc[i].duplicated().any():
            print('\n==========ERROR==========')
            print(f"Row of the element {df_domain_selection.index[i]} in df_domain_selection has duplicated values. Please select only one domain for each type of energy of the component")
            sys.exit()

    #checking if elements have more domains defines than they are design to produce/consume
    df_temp = df_domain_selection[list_columns].copy()
    for i in range(len(df_temp)):
        if sum(1 for s in df_temp.iloc[i] if s != 0) > len(eval(df_domain_selection['component_domains'].iloc[i])):
            print('\n==========ERROR==========')
            print(f"Row of the element {df_domain_selection.index[i]} has more domains defined than what the component is designed to handle.")
            sys.exit()

    #getting inputs and seprating element into sources and loads of the selected domains
    list_source_final = []
    list_load_final = []
    for i in df_domain.index:
        domain = df_domain.loc[i,'domain_names']
        list_source_domain = []
        list_load_domain = []
        for j in df_domain_selection.index:
            list_component_domain = eval(df_domain_selection.loc[j,'component_domains'])
            list_component_source_domain = eval(df_domain_selection.loc[j,'component_source_domains'])
            list_component_load_domain = eval(df_domain_selection.loc[j,'component_load_domains'])
            for m,n in enumerate(list_columns):
                if df_domain_selection.loc[j,n] == domain:
                    if list_component_domain[m] in list_component_source_domain:
                        list_source_domain.append(j)
                    if list_component_domain[m] in list_component_load_domain:
                        list_load_domain.append(j) 

        list_source_final.append(list_source_domain)
        list_load_final.append(list_load_domain)

    #creating and writing connection tables for each domain
    for i in range(len(df_domain)):
        df_temp = pd.DataFrame(0,index = list_load_final[i], columns = list_source_final[i])
        name_sheet = 'connect_domain_' + df_domain['domain_names'].iloc[i]
        write_excel(df_temp, control.path_input, name_sheet,'input.xlsx',True)


    #creating and writing revenue tables for each domain
    for i in range(len(df_domain)):
        df_temp = pd.DataFrame(0,index = list_load_final[i], columns = list_source_final[i])
        name_sheet = 'revenue_domain_' + df_domain['domain_names'].iloc[i]
        write_excel(df_temp, control.path_input, name_sheet,'input.xlsx',True)

    #creating and writing stock tables for each domain
    for i in range(len(df_domain)):
        df_temp = pd.DataFrame(0,index = list_load_final[i], columns = list_source_final[i])
        df_temp = df_temp.filter(like = 'net', axis = 0)
        name_sheet = 'stock_domain_' + df_domain['domain_names'].iloc[i]
        write_excel(df_temp, control.path_input, name_sheet,'input.xlsx',True)

    return 

def create_connection_equations(df_domains,control):
    list_attr_classes = []
    list_expressions = []
    list_con_variables = []
    list_connection_matrices = []
    for i in df_domains.index:
        domain_name = df_domains.loc[i,'domain_names']
        sheet_name = 'connect_domain_' + domain_name
        df_connect = pd.read_excel(control.path_input + control.name_file, sheet_name = sheet_name,index_col=0)
        df_connect.index.name = None

        list_attr_classes = list_attr_classes + df_connect.index.to_list()
        list_attr_classes = list_attr_classes + df_connect.columns.to_list()

        #building variable names inside connection
        for j in df_connect.index: 
            for k in df_connect.columns:
                if df_connect.loc[j,k] != 0:
                    if j == k:
                        print('\n==========ERROR==========')
                        print('Energy cannot be transferred from ' + j + ' to ' + k +'\n')
                        sys.exit()
                    else:
                        df_connect.loc[j,k] = domain_name + '_' + k + '_' + j

        #creating variable names derived from columns of conection matrix
        list_sub = [domain_name + '_from_' + s for s in df_connect.columns]
        df_connect.columns = list_sub

        #creating variable names derived from index of conection matrix
        list_sub = [domain_name + '_to_' + s for s in df_connect.index]
        list_sub = [s.replace(domain_name + '_to_demand','param_P_to_demand') for s in list_sub]
        list_sub = [s.replace(domain_name + '_to_charging_station','param_P_to_charging_station') for s in list_sub]
        df_connect.index = list_sub

        df_connect.to_excel(control.path_output + 'df_connect_' + domain_name + '.xlsx')

        # creating equations in the 'to' direction of connection matrix
        for j in df_connect.index:
            if j.startswith('param_P_to_demand'): #doing adjustment to include penalty power if demand cannot be covered.
                name_demand = j.replace('param_P_to_','')
                list_exp_partial = 'model.' + j + '[t] == model.'+ name_demand + '_P_extra[t]'
            else:
                list_exp_partial = 'model.' + j + '[t] == 0'

            for k in df_connect.columns:
                if df_connect.loc[j,k] != 0:
                    list_exp_partial = list_exp_partial + ' + model.' + df_connect.loc[j,k] + '[t]'
            list_expressions.append(list_exp_partial)
        
        # creating equations in the 'from' direction of connection matrix
        for j in df_connect.columns:
            list_exp_partial = 'model.' + j + '[t] == 0'
            for k in df_connect.index:
                if df_connect.loc[k,j] != 0:
                    list_exp_partial = list_exp_partial + ' + model.' + df_connect.loc[k,j] + '[t]'
            list_expressions.append(list_exp_partial)

        # creating list with variables that need to be created in the abstract model ELECTRIC
        list_con_variables = list_con_variables + df_connect.columns.to_list()
        list_con_variables = list_con_variables + df_connect.index.to_list()
        for j in df_connect.index:
            for k in df_connect.columns:
                if df_connect.loc[j,k] != 0:
                    list_con_variables.append(df_connect.loc[j,k])
        list_con_variables = list(set(list_con_variables))

        list_connection_matrices.append(df_connect)

    print('\n---------------This is the list_attr_classes---------------')
    print(list_attr_classes)
    print('\n---------------This is the list_expressions---------------')
    print(list_expressions)
    print('\n---------------This is the list_con_variables---------------')
    print(list_con_variables)

    return list_connection_matrices, list_expressions, list_con_variables, list_attr_classes

def create_revenue_and_stock_equations(df_domains,df_input_other,control):
    list_expressions_rev = []
    list_variables_rev = []
    list_parameters_rev = []
    list_parameters_rev_value = []
    list_correl_elements = []
    for i in df_domains.index:
        domain_name = df_domains.loc[i,'domain_names']
        df_revenue = pd.read_excel(control.path_input + control.name_file, sheet_name = 'revenue_domain_' + domain_name ,index_col=0)
        df_revenue.index.name = None
        df_connect = pd.read_excel(control.path_input + control.name_file, sheet_name = 'connect_domain_' + domain_name ,index_col = 0)
        df_revenue.index.name = None
        df_stock = pd.read_excel(control.path_input + control.name_file, sheet_name = 'stock_domain_' + domain_name, index_col = 0)

        #updating df_connect again with respective energy flows
        for j in df_connect.index:
            for k in df_connect.columns:
                df_connect.loc[j,k] = domain_name + '_' + k + '_' + j

        #creating expressions for revenue from incentives
        for j in df_connect.columns:
            for k in df_connect.index:
                if df_connect.loc[k,j] != 0:
                    if df_revenue.loc[k,j] != 0:
                        list_correl_elements.append(j) 
                        name_parameter = ('param_rev_' + domain_name + '_' + j + '_' + k)
                        name_variable = ('rev_'+ domain_name +'_' + j + '_' + k)
                        name_energy_flow = df_connect.loc[k,j]
                        list_parameters_rev.append(name_parameter)
                        list_parameters_rev_value.append(df_revenue.loc[k,j])
                        list_variables_rev.append(name_variable)
                        list_expressions_rev.append(F"model.{name_variable}[t] == model.{name_parameter} * model.{name_energy_flow}[t] * model.time_step")

        #creating expressions for revenue from selling energy to the energy stock exchange
        df_connect_stock = df_connect.filter(like = 'net', axis = 0)
        for j in df_connect_stock.columns:
            for k in df_connect_stock.index:
                if df_connect_stock.loc[k,j] != 0:
                    if df_stock.loc[k,j] != 0:
                        name_variable = ('stock_'+ domain_name +'_' + j + '_' + k)
                        name_energy_flow = df_connect_stock.loc[k,j]
                        list_variables_rev.append(name_variable)
                        list_expressions_rev.append(F"model.{name_variable}[t] == model.param_{k}_stock_price_electric[t] * model.{name_energy_flow}[t] * model.time_step")

        #list with expression of total revenues
        list_revenue_total = ['model.total_revenue[t] ==']
        if list_expressions_rev == []:
            list_revenue_total = list_revenue_total  + '0'
        else:
            for j in list_variables_rev:
                list_revenue_total[-1] = list_revenue_total[-1] + ' + model.' + j + '[t]'

    #saving values of parameters to df_input_other
    df_parameters = pd.DataFrame({'Parameter':list_parameters_rev,
                                'Value':list_parameters_rev_value})
    df_input_other = pd.concat([df_input_other, df_parameters], ignore_index = True)

    df_input_other.to_excel(control.path_output + 'df_input_other_test.xlsx',index = False)

    print('\n--------------------------This is list_expressions_rev--------------------------')
    print(list_expressions_rev)

    print('\n--------------------------This is list_variables_rev--------------------------')
    print(list_variables_rev)

    print('\n--------------------------This is list_revenue_total--------------------------')
    print(list_revenue_total)

    print('\n--------------------------This is list_correl_elements--------------------------')
    print(list_correl_elements)

    return df_input_other, list_expressions_rev, list_variables_rev, list_parameters_rev, list_parameters_rev_value, list_revenue_total, list_correl_elements



def objective_expression_creator(df_aux):
    #total investment cost expression
    list_investment_costs_total = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        method = getattr(classes,element_type)
        if hasattr(method,'constraint_investment_costs'):
            if list_investment_costs_total == []:
                list_investment_costs_total.append('model.total_investment_costs[t] == ' + ' model.' + element + '_inv_costs[t]')
            else:
                list_investment_costs_total[-1] = list_investment_costs_total[-1] + ' + model.'+ element + '_inv_costs[t]'

    print('\n======================================This is the total investment costs')
    print(list_investment_costs_total)


    #total operation costs expression
    list_operation_costs_total = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        method = getattr(classes,element_type)
        if hasattr(method,'constraint_operation_costs'):
            if list_operation_costs_total == [] :
                list_operation_costs_total.append('model.total_operation_costs[t] == ' + ' model.' + element + '_op_costs[t]')
            else:
                list_operation_costs_total[-1] = list_operation_costs_total[-1] + ' + model.'+ element + '_op_costs[t]'

    print('\n======================================This is the total operation costs')
    print(list_operation_costs_total)


    #total emissions expressions
    list_emissions_total = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        method = getattr(classes,element_type)
        if hasattr(method,'constraint_emissions'):
            if list_emissions_total == [] :
                list_emissions_total.append('model.total_emissions[t] == ' + 'model.' + element +'_emissions[t]')
            else:
                list_emissions_total[-1] = list_emissions_total[-1] + ' + model.' + element +'_emissions[t]'

    print('\n======================================This is the total emission costs')
    print(list_emissions_total)

    return list_operation_costs_total, list_investment_costs_total, list_emissions_total

# def objective_constraint_creator(df_aux): # this function creates the constraints for the objective function to work

#     # creating constraints to calculate bouhgt and sold energy from the grid.
#     list_buy_constraint = ['model.total_buy[t] == '] 
#     list_sell_constraint = ['model.total_sell[t] == ']

#     df_temp = df_aux[df_aux['type'] == 'net' ].reset_index(drop = True)
    
#     for i in df_temp.index:
#         element = df_temp['element'].iloc[i]
#         if i == 0:
#             list_buy_constraint[-1] = list_buy_constraint[-1] + ' model.' + element + '_buy_electric[t]' + ' + model.' + element + '_buy_thermal[t]'
#             list_sell_constraint[-1] = list_sell_constraint[-1] + ' model.' + element + '_sell_electric[t]' + ' + model.' + element + '_sell_thermal[t]'
#         else:
#             list_buy_constraint[-1] = list_buy_constraint[-1] + ' + model.' + element + '_buy_electric[t]' + ' + model.' + element + '_buy_thermal[t]'
#             list_sell_constraint[-1] = list_sell_constraint[-1] + ' + model.' + element + '_sell_electric[t]' + ' + model.' + element + '_sell_thermal[t]'
    
#     list_objective_constraints = list_buy_constraint + list_sell_constraint


#     list_operation_costs_total = []
#     for i in df_aux.index:
#         element = df_aux['element'].iloc[i]
#         element_type = df_aux['type'].iloc[i]
#         method = getattr(classes,element_type)
#         if hasattr(method,'constraint_operation_costs'):
#             if list_operation_costs_total == [] :
#                 list_operation_costs_total.append('model.total_operation_costs[t] == ' + ' model.' + element + '_op_costs[t]')
#             else:
#                 list_operation_costs_total[-1] = list_operation_costs_total[-1] + '+ model.'+ element + '_op_costs[t]'

#     list_objective_constraints = list_objective_constraints + list_operation_costs_total

#     list_emissions_constraint = []
#     for i in df_aux.index:
#         element = df_aux['element'].iloc[i]
#         element_type = df_aux['type'].iloc[i]
#         method = getattr(classes,element_type)
#         if hasattr(method,'constraint_emissions'):
#             if list_emissions_constraint == [] :
#                 list_emissions_constraint.append('model.total_emissions[t] == ' + 'model. ' + element +'_emissions[t]')
#             else:
#                 list_emissions_constraint[-1] = list_emissions_constraint[-1] + ' + model. ' + element +'_emissions[t]'
    
#     # print('\nEmission Constriants') 
#     # print(list_emissions_constraint)

#     list_objective_constraints = list_objective_constraints + list_emissions_constraint

#     return list_objective_constraints

def organize_output_columns(df_variable_values,df_aux):
    list_columns = df_variable_values.columns
    list_temp = []
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        for j in list_columns:
            if '_'+ element +'_' in j:
                list_temp.append(j)
            elif element +'_' in j:
                list_temp.append(j)
            elif 'from_'+ element in j:
                list_temp.append(j)
            elif 'to_' + element in j:
                list_temp.append(j)

    complimentary_list = list(set(list_columns) - set(list_temp))
    list_temp = list_temp + complimentary_list
    list_temp.remove('TimeStep')
    list_temp.insert(0,'TimeStep')
    df_variable_values = df_variable_values.reindex(columns = list_temp)

    return df_variable_values

def breaking_dataframe(df,horizon,saved_position):
    list_split = []
    for i in range(0,len(df),saved_position):
        # print('\n------------'+str(i))
        if i == 0:
            list_split.append(df.iloc[i : i + horizon])
        else:
            list_split.append(df.iloc[i-1 : i+horizon])
        # print(list_split[-1])
    return list_split

def save_variables_last_time_step(df_input_others,df_variables_last_time_step):
    for j in df_variables_last_time_step.index:
        if df_variables_last_time_step['Parameter'].iloc[j] in df_input_others['Parameter'].tolist():
            df_input_others.loc[df_input_others['Parameter'] == df_variables_last_time_step['Parameter'].iloc[j], 'Value'] = df_variables_last_time_step['Value'].iloc[j]
        else:
            row_to_append = df_variables_last_time_step.loc[j]
            df_input_others = df_input_others.append(row_to_append, ignore_index = True)
            
    return df_input_others

def emissions_analysis(control):

    warnings.filterwarnings("ignore")

    path_output = control.path_output

    df_final = pd.read_excel(path_output + 'df_final.xlsx')
    reference_date = control.reference_date

    df_final['Date'] = reference_date + pd.to_timedelta(df_final['TimeStep'], unit = 'h')
    df_final.set_index('Date', inplace = True)

    df_final['Year'] = df_final.index.year.astype(int)
    df_final['Week'] = df_final.index.week.astype(int)
    df_final['Year Month'] = df_final.index.strftime('%Y-%m')
    df_final['Year Week'] = df_final.index.strftime('%Y-%W').astype(str)
    df_final['Year Week'] = [s.replace('-', ' - wk ') for s in df_final['Year Week']]

    weekly_sum_df = df_final.groupby('Year Week').sum()
    monthly_sum_df = df_final.groupby('Year Month').sum()
    annual_sum_df = df_final.groupby('Year').sum()

    # Create a new Excel workbook
    wb = Workbook()

    # Select the active sheet (you can also create a new one if needed)
    sheet = wb.active

    #some formats
    thousands_separator_format = NamedStyle(name="thousands_separator_format")
    thousands_separator_format.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'

    centered_style = NamedStyle(name="centered_style")
    centered_style.alignment = Alignment(horizontal = "center", vertical = "center")

    normal_style = NamedStyle(name="normal_style")
    normal_style.alignment = Alignment(horizontal = "left", vertical = "center", indent = 0)

    normal_style_with_indent = NamedStyle(name="normal_style_with_indent")
    normal_style_with_indent.alignment = Alignment(horizontal = "left", vertical = "center", indent = 2)


    # function for formatting cells
    def formatting_cells(minRow, maxRow, minCol, maxCol, colorCode, boldOption, fontOption, sizeOption, styleOption):
        for row in sheet.iter_rows(min_row = minRow, max_row = maxRow, min_col = minCol, max_col = maxCol):
            for cell in row:
                cell.style = styleOption
                cell.fill = PatternFill(start_color = colorCode, end_color=colorCode, fill_type="solid")
                cell.font = Font(bold = boldOption, size = sizeOption, name = fontOption)
        return


    def conditional_formatting_cells(minRow, maxRow, minCol, maxCol, colorCode, boldOption, fontOption, sizeOption, styleOption,fontColor):
        for cell in sheet.iter_rows(min_row = minRow, max_row = maxRow, min_col = minCol, max_col = maxCol):
            for c in cell:
                c.style = styleOption
                c.fill = PatternFill(start_color = colorCode, end_color = colorCode, fill_type="solid")
                if c.value >= 0:
                    c.font = Font(bold = boldOption, size = sizeOption, name = fontOption)
                else:
                    c.font = Font(bold = boldOption, size = sizeOption, name = fontOption, color = fontColor)
        return

    def set_column_width(width, index):
        column_index = index
        column_letter = utils.get_column_letter(column_index)
        column_dimension = sheet.column_dimensions[column_letter]
        column_dimension.width = width
        return

    # function for inserting values

    # Inserting values and formating cells from the top of the sheet to down
    # --------------------------------------------------------------------------
    # region top strips

    sheet.cell(row = 2, column = 2, value = 'Emissions Analysis - Calculations (EoP)')
    sheet.cell(row = 3, column = 2, value = 'Units: kg CO2')

    formatting_cells(minRow = 2, maxRow = 2, minCol = 1, maxCol = 1000, colorCode = "009EE3", 
                    boldOption = True, sizeOption = 11, fontOption = 'Arial', styleOption= normal_style)

    formatting_cells(minRow = 3, maxRow = 3, minCol = 1, maxCol = 1000, colorCode = "93D3F7", 
                    boldOption = False, sizeOption = 11, fontOption = 'Arial', styleOption= normal_style)

    # endregion
    # --------------------------------------------------------------------------
    # region inserting operation costs

    col_index_yearly_data = 5
    col_index_monthly_data = 5 + len(annual_sum_df) + 1
    col_index_weekly_data = 5 + len(annual_sum_df) + 1 + len(monthly_sum_df) + 1

    # operation costs title
    sheet.cell(row = 7, column = 3, value = '1 - Emissions')

    #formating row operation costs
    formatting_cells(minRow = 7, maxRow = 7, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)

    # Year, month and weeks title
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_yearly_data, maxCol = col_index_yearly_data + len(annual_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_monthly_data, maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_weekly_data, maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)


    # inserting and formatting ANNUAL DATA
    list_columns = [s for s in annual_sum_df if 'emissions' in s and 'total' not in s]
    annual_emissions_df = annual_sum_df[list_columns]
    # annual_emissions_df *= -1

    #inserting title of row columns
    for idx, value in enumerate(annual_emissions_df.columns, start = 8):
        sheet.cell(row = idx, column = 3, value = value)

    #inserting years titles
    for idx, value in enumerate(annual_emissions_df.index, start = 5):
        sheet.cell(row = 5, column = idx, value = value)

    #inserting values of yearly data
    for row_idx, row in enumerate(annual_emissions_df.iterrows(), start = 5):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    #formatting titles of row columns
    formatting_cells(minRow = 8, maxRow =  8 + len(annual_emissions_df.columns) - 1, minCol = 1, maxCol = 3, colorCode = "D4EEFC", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption= normal_style_with_indent)

    #formatting numbers of yearly values
    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(annual_emissions_df.columns) - 1, minCol = 5, maxCol = 5 + len(annual_sum_df) - 1, 
                                    colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "00B050")

    # formatting annual sum for the years
    list_annual_sum = annual_emissions_df.sum(axis =1)
    for col_idx, value in enumerate(list_annual_sum, start = 5):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = 5, maxCol = 5 + len(list_annual_sum) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "00B050")


    # inserting and formatting MONTHLY DATA
    list_columns = [s for s in monthly_sum_df if 'emissions' in s and 'total' not in s]
    monthly_emissions_df = monthly_sum_df[list_columns]
    # monthly_emissions_df *= -1 

    #inserting month titles
    for idx, value in enumerate(monthly_emissions_df.index, start = col_index_monthly_data):
        sheet.cell(row = 5, column = idx, value = value)

    for row_idx, row in enumerate(monthly_emissions_df.iterrows(), start = col_index_monthly_data):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(monthly_emissions_df.columns) - 1, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")

    # formatting annual sum for the months
    list_monthly_sum = monthly_emissions_df.sum(axis =1)
    for col_idx, value in enumerate(list_monthly_sum, start = col_index_monthly_data):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting monthly sum for the months
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")


    # inserting and formatting WEEKLY DATA
    list_columns = [s for s in weekly_sum_df if 'emissions' in s and 'total' not in s]
    weekly_emissions_df = weekly_sum_df[list_columns]
    # weekly_costs_df *= -1

    #inserting week titles
    for idx, value in enumerate(weekly_emissions_df.index, start = col_index_weekly_data):
        sheet.cell(row = 5, column = idx, value = value)

    for row_idx, row in enumerate(weekly_emissions_df.iterrows(), start = col_index_weekly_data):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(weekly_emissions_df.columns) - 1, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")

    # formatting annual sum for the weeks
    list_weekly_sum = weekly_emissions_df.sum(axis =1)
    for col_idx, value in enumerate(list_weekly_sum, start = col_index_weekly_data):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting monthly sum for the weeks
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")


    # endregion
    # --------------------------------------------------------------------------
    # region other adjustments

    set_column_width(23, 3)
    set_column_width(2, 4)
    set_column_width(2, col_index_monthly_data - 1)
    set_column_width(2, col_index_weekly_data -1)

    for i in range(col_index_weekly_data, col_index_weekly_data + len(weekly_emissions_df)):
        set_column_width(13, i)

    # endregion
    # --------------------------------------------------------------------------
    # region other adjustments

    annual_result  = np.array(list_annual_sum) 
    monthly_result  = np.array(list_monthly_sum) 
    weekly_result  = np.array(list_weekly_sum) 


    # now inserting the cumulated sum of the cash flows
    row_index_total_cumulated = 7 + len(annual_emissions_df.columns) + 1 + 1

    # operation costs title
    sheet.cell(row = row_index_total_cumulated, column = 3, value = '2 - Emissions accumulated')

    #formating row operation costs
    formatting_cells(minRow = row_index_total_cumulated, maxRow = row_index_total_cumulated, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)
    
    
    # inserting annual result for the years
    for col_idx, value in enumerate(np.cumsum(annual_result), start = 5):
        sheet.cell(row = row_index_total_cumulated, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = row_index_total_cumulated, maxRow  = row_index_total_cumulated, minCol = 5, maxCol = 5 + len(annual_result) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "00B050")

    # inserting annual sum for the months
    for col_idx, value in enumerate(np.cumsum(monthly_result), start = col_index_monthly_data):
        sheet.cell(row =  row_index_total_cumulated, column = col_idx, value = value)

    # formatting monthly sum for the years
    conditional_formatting_cells(minRow =  row_index_total_cumulated, maxRow  =  row_index_total_cumulated, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(monthly_result) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")

    # inserting total sum for the weeks
    for col_idx, value in enumerate(np.cumsum(weekly_result), start = col_index_weekly_data):
        sheet.cell(row = row_index_total_cumulated, column = col_idx, value = value)

    # formatting total sum for the weeks
    conditional_formatting_cells(minRow = row_index_total_cumulated, maxRow  = row_index_total_cumulated, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_result) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "00B050")

    #endregion

    # Hide gridlines
    sheet.sheet_view.showGridLines = False
    mycell = sheet['D6']
    sheet.freeze_panes = mycell

    # Save the Excel file
    wb.save(path_output + "emission_analysis.xlsx")

def charts_generator(control,df_aux,df_domains):
    path_output = control.path_output
    path_charts = control.path_charts
    reference_date = control.reference_date

    df_final = pd.read_excel(path_output + 'df_final.xlsx')
    df_final['Date'] = reference_date + pd.to_timedelta(df_final['TimeStep'], unit = 'h')
    if control.time_span < 100:
        df_final = df_final[:control.time_span]
    else:
        df_final = df_final[:100]

    x_axis = df_final['TimeStep'].tolist()
    figure_size = (15,8)


    #--------------NOVO
    for i in df_aux.index:
        element = df_aux['element'].iloc[i]
        element_type = df_aux['type'].iloc[i]
        for j in df_domains.index:
            domain_name = df_domains.loc[j,'domain_names']
            print('\n-------------')
            print(domain_name)
            folder_path = path_charts + '/' + element + '/'

            try:
                os.mkdir(folder_path)
            except FileExistsError:
                print(f"Folder '{element}' already exists at {folder_path}")
            except Exception as e:
                print(f"An error occurred: {str(e)}")

            list_connections_domain = [s for s in df_final.columns if '_' + element + '_' in s and domain_name in s and 'rev' not in s and 'stock' not in s]

            plt.figure(figsize = figure_size)
            bottom = np.zeros(len(x_axis))
            if len(list_connections_domain) > 0:
                for k in list_connections_domain:
                    plt.bar(x_axis, df_final[k], bottom = bottom)
                    bottom += df_final[k]
                plt.xlabel('Time [hs]')
                plt.ylabel('Power [kW]')
                plt.title(f"{domain_name} - {element}")
                plt.legend(list_connections_domain)
                plt.savefig(folder_path + domain_name + ' - ' + element + '.png')
                plt.close()

            if element_type == 'bat':
                plt.figure(figsize = figure_size)
                plt.bar(x_axis, df_final[element + '_energy'])
                plt.xlabel('Time [hs]')
                plt.ylabel('Energy [kWh]')
                plt.title(f"Energy level - {element}")
                plt.legend(element)
                plt.savefig(folder_path + 'E - ' + element +'.png')
                plt.close()

                list_charging_bat = [s for s in df_final.columns if s.endswith(element) and 'to' not in s and 'from' not in s]
                plt.figure(figsize = figure_size)
                bottom = np.zeros(len(x_axis))
                if len(list_charging_bat) > 0:
                    for j in list_charging_bat:
                        plt.bar(x_axis, df_final[j], bottom = bottom)
                        bottom += df_final[j]
                    plt.xlabel('Time [hs]')
                    plt.ylabel('Power [kW]')
                    plt.title(f"Charging power distribution - {element}")
                    plt.legend(list_charging_bat)
                    plt.savefig(folder_path + 'Pin - ' + element + '.png')
                    plt.close()

            elif element_type == 'bat_with_aging':
                plt.figure(figsize = figure_size)
                plt.bar(x_axis, df_final[element + '_SOC'])
                plt.xlabel('Time [hs]')
                plt.ylabel('State of charge [-]')
                plt.title(f"State of charge - {element}")
                plt.legend(element)
                plt.savefig(folder_path + 'SoC - ' + element +'.png')
                plt.close()

                plt.figure(figsize = figure_size)
                plt.plot(x_axis, df_final[element + '_SOC_max'])
                plt.xlabel('Time [hs]')
                plt.ylabel('max state of charge [-]')
                plt.title(f"Effect of aging on max state of charge - {element}")
                plt.legend(element)
                plt.savefig(folder_path + 'max SoC - ' + element +'.png')
                plt.close()

                list_charging_bat = [s for s in df_final.columns if s.endswith(element) and 'to' not in s and 'from' not in s]
                plt.figure(figsize = figure_size)
                bottom = np.zeros(len(x_axis))
                if len(list_charging_bat) > 0:
                    for j in list_charging_bat:
                        plt.bar(x_axis, df_final[j], bottom = bottom)
                        bottom += df_final[j]
                    plt.xlabel('Time [hs]')
                    plt.ylabel('Power [kW]')
                    plt.title(f"Electric charging power distribution - {element}")
                    plt.legend(list_charging_bat)
                    plt.savefig(folder_path + 'Pin - ' + element + '.png')
                    plt.close()


    #--------------NOVO

    # for i in df_aux.index:
    #     element = df_aux['element'].iloc[i]
    #     element_type =  df_aux['type'].iloc[i]

    #     list_connections_electric = [s for s in df_final.columns if '_' + element + '_' in s and 'P_' in s and 'rev' not in s and 'stock' not in s]
    #     list_connections_thermal = [s for s in df_final.columns if '_' + element + '_' in s and 'Q_' in s and 'rev' not in s and 'stock' not in s]

    #     folder_path = path_charts + '/' + element + '/'

    #     try:
    #         os.mkdir(folder_path)
    #     except FileExistsError:
    #         print(f"Folder '{element}' already exists at {folder_path}")
    #     except Exception as e:
    #         print(f"An error occurred: {str(e)}")

    #     plt.figure(figsize = figure_size)
    #     bottom = np.zeros(len(x_axis))
    #     if len(list_connections_electric) > 0:
    #         for j in list_connections_electric:
    #             plt.bar(x_axis, df_final[j], bottom = bottom)
    #             bottom += df_final[j]
    #         plt.xlabel('Time [hs]')
    #         plt.ylabel('Power [kW]')
    #         plt.title(f"Electric power distribution - {element}")
    #         plt.legend(list_connections_electric)
    #         plt.savefig(folder_path + 'P - ' + element + '.png')
    #         plt.close()

    #     plt.figure(figsize = figure_size)
    #     bottom = np.zeros(len(x_axis))
    #     if len(list_connections_thermal) > 0:
    #         for j in list_connections_thermal:
    #             plt.bar(x_axis, df_final[j],bottom = bottom)
    #             bottom += df_final[j]
    #         plt.xlabel('Time [hs]')
    #         plt.ylabel('Power [kW]')
    #         plt.title(f"Thermal power distribution - {element}")
    #         plt.legend(list_connections_thermal)
    #         plt.savefig(folder_path +'Q - ' + element + '.png')
    #         plt.close()

    #     if element_type == 'bat':
    #         plt.figure(figsize = figure_size)
    #         plt.bar(x_axis, df_final[element + '_energy'])
    #         plt.xlabel('Time [hs]')
    #         plt.ylabel('Energy [kWh]')
    #         plt.title(f"Energy level - {element}")
    #         plt.legend(element)
    #         plt.savefig(folder_path + 'E - ' + element +'.png')
    #         plt.close()

    #         list_charging_bat = [s for s in df_final.columns if s.endswith(element) and 'to' not in s and 'from' not in s]
    #         plt.figure(figsize = figure_size)
    #         bottom = np.zeros(len(x_axis))
    #         if len(list_charging_bat) > 0:
    #             for j in list_charging_bat:
    #                 plt.bar(x_axis, df_final[j], bottom = bottom)
    #                 bottom += df_final[j]
    #             plt.xlabel('Time [hs]')
    #             plt.ylabel('Power [kW]')
    #             plt.title(f"Electric charging power distribution - {element}")
    #             plt.legend(list_charging_bat)
    #             plt.savefig(folder_path + 'Pin - ' + element + '.png')
    #             plt.close()


    #     elif element_type == 'bat_with_aging':
    #         plt.figure(figsize = figure_size)
    #         plt.bar(x_axis, df_final[element + '_SOC'])
    #         plt.xlabel('Time [hs]')
    #         plt.ylabel('State of charge [-]')
    #         plt.title(f"State of charge - {element}")
    #         plt.legend(element)
    #         plt.savefig(folder_path + 'SoC - ' + element +'.png')
    #         plt.close()

    #         plt.figure(figsize = figure_size)
    #         plt.plot(x_axis, df_final[element + '_SOC_max'])
    #         plt.xlabel('Time [hs]')
    #         plt.ylabel('max state of charge [-]')
    #         plt.title(f"Effect of aging on max state of charge - {element}")
    #         plt.legend(element)
    #         plt.savefig(folder_path + 'max SoC - ' + element +'.png')
    #         plt.close()

    #         list_charging_bat = [s for s in df_final.columns if s.endswith(element) and 'to' not in s and 'from' not in s]
    #         plt.figure(figsize = figure_size)
    #         bottom = np.zeros(len(x_axis))
    #         if len(list_charging_bat) > 0:
    #             for j in list_charging_bat:
    #                 plt.bar(x_axis, df_final[j], bottom = bottom)
    #                 bottom += df_final[j]
    #             plt.xlabel('Time [hs]')
    #             plt.ylabel('Power [kW]')
    #             plt.title(f"Electric charging power distribution - {element}")
    #             plt.legend(list_charging_bat)
    #             plt.savefig(folder_path + 'Pin - ' + element + '.png')
    #             plt.close()

def financial_analysis(control):

    warnings.filterwarnings("ignore")

    path_output = control.path_output

    df_final = pd.read_excel(path_output + 'df_final.xlsx')
    reference_date = control.reference_date

    df_final['Date'] = reference_date + pd.to_timedelta(df_final['TimeStep'], unit = 'h')
    df_final.set_index('Date', inplace = True)

    df_final['Year'] = df_final.index.year.astype(int)
    df_final['Week'] = df_final.index.week.astype(int)
    df_final['Year Month'] = df_final.index.strftime('%Y-%m')
    df_final['Year Week'] = df_final.index.strftime('%Y-%W').astype(str)
    df_final['Year Week'] = [ s.replace('-', ' - wk ') for s in df_final['Year Week']]

    weekly_sum_df = df_final.groupby('Year Week').sum()
    monthly_sum_df = df_final.groupby('Year Month').sum()
    annual_sum_df = df_final.groupby('Year').sum()

    # Create a new Excel workbook
    wb = Workbook()

    # Select the active sheet (you can also create a new one if needed)
    sheet = wb.active

    #some formats
    thousands_separator_format = NamedStyle(name="thousands_separator_format")
    thousands_separator_format.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'

    centered_style = NamedStyle(name="centered_style")
    centered_style.alignment = Alignment(horizontal = "center", vertical = "center")

    normal_style = NamedStyle(name="normal_style")
    normal_style.alignment = Alignment(horizontal = "left", vertical = "center", indent = 0)

    normal_style_with_indent = NamedStyle(name="normal_style_with_indent")
    normal_style_with_indent.alignment = Alignment(horizontal = "left", vertical = "center", indent = 2)

    #defining indexes for columns
    col_index_yearly_data = 5
    col_index_monthly_data = 5 + len(annual_sum_df) + 1
    col_index_weekly_data = 5 + len(annual_sum_df) + 1 + len(monthly_sum_df) + 1

    # function for formatting cells
    def formatting_cells(minRow, maxRow, minCol, maxCol, colorCode, boldOption, fontOption, sizeOption, styleOption):
        for row in sheet.iter_rows(min_row = minRow, max_row = maxRow, min_col = minCol, max_col = maxCol):
            for cell in row:
                cell.style = styleOption
                cell.fill = PatternFill(start_color = colorCode, end_color=colorCode, fill_type="solid")
                cell.font = Font(bold = boldOption, size = sizeOption, name = fontOption)
        return

    def conditional_formatting_cells(minRow, maxRow, minCol, maxCol, colorCode, boldOption, fontOption, sizeOption, styleOption,fontColor):
        for cell in sheet.iter_rows(min_row = minRow, max_row = maxRow, min_col = minCol, max_col = maxCol):
            for c in cell:
                c.style = styleOption
                c.fill = PatternFill(start_color = colorCode, end_color = colorCode, fill_type="solid")
                if c.value >= 0:
                    c.font = Font(bold = boldOption, size = sizeOption, name = fontOption)
                else:
                    c.font = Font(bold = boldOption, size = sizeOption, name = fontOption, color = fontColor)
        return

    def set_column_width(width, index):
        column_index = index
        column_letter = utils.get_column_letter(column_index)
        column_dimension = sheet.column_dimensions[column_letter]
        column_dimension.width = width
        return


    # Inserting values and formating cells from the top of the sheet to dowm
    # --------------------------------------------------------------------------
    # region top strips

    sheet.cell(row = 2, column = 2, value = 'Financial Analysis - Calculations (EoP)')
    sheet.cell(row = 3, column = 2, value = 'Units: €')

    formatting_cells(minRow = 2, maxRow = 2, minCol = 1, maxCol = 1000, colorCode = "009EE3", 
                    boldOption = True, sizeOption = 11, fontOption = 'Arial', styleOption= normal_style)

    formatting_cells(minRow = 3, maxRow = 3, minCol = 1, maxCol = 1000, colorCode = "93D3F7", 
                    boldOption = False, sizeOption = 11, fontOption = 'Arial', styleOption= normal_style)
    
    # endregion
    # ---------------------------------------------------------------------------------------------------------------------------------------
    # region inserting REVENUE VALUES

    # operation costs title
    sheet.cell(row = 7, column = 3, value = '1 - Revenue')

    #formating row operation costs
    formatting_cells(minRow = 7, maxRow = 7, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)
    
    # formatting year, month and weeks 
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_yearly_data, maxCol = col_index_yearly_data + len(annual_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_monthly_data, maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)
    formatting_cells(minRow = 5, maxRow = 5, minCol = col_index_weekly_data, maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, colorCode = "D9D9D9", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = centered_style)
    
    # inserting and formatting ANNUAL DATA
    list_columns = [s for s in annual_sum_df if '_rev' in s or 'rev_' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    annual_revenue_df = annual_sum_df[list_columns]

    #inserting title of row titles
    for idx, value in enumerate(annual_revenue_df.columns, start = 8):
        sheet.cell(row = idx, column = 3, value = value)

    #inserting years titles
    for idx, value in enumerate(annual_revenue_df.index, start = 5):
        sheet.cell(row = 5, column = idx, value = value)

    #inserting values of yearly data
    for row_idx, row in enumerate(annual_revenue_df.iterrows(), start = 5):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)


    #formatting titles of row columns
    formatting_cells(minRow = 8, maxRow =  8 + len(annual_revenue_df.columns) - 1, minCol = 1, maxCol = 3, colorCode = "D4EEFC", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption= normal_style_with_indent)

    #formatting numbers of yearly values
    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(annual_revenue_df.columns) - 1, minCol = 5, maxCol = 5 + len(annual_sum_df) - 1, 
                                    colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")

    # formatting annual sum for the years
    list_annual_sum_revenue = annual_revenue_df.sum(axis =1)
    for col_idx, value in enumerate(list_annual_sum_revenue, start = 5):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = 5, maxCol = 5 + len(list_annual_sum_revenue) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")
    





    # inserting and formatting MONTHLY DATA
    list_columns = [s for s in monthly_sum_df if 'rev_' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    monthly_revenue_df = monthly_sum_df[list_columns]

    #inserting month titles
    for idx, value in enumerate(monthly_revenue_df.index, start = col_index_monthly_data):
        sheet.cell(row = 5, column = idx, value = value)

    for row_idx, row in enumerate(monthly_revenue_df.iterrows(), start = col_index_monthly_data):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(annual_revenue_df.columns) - 1, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")
    

   # formatting annual sum for the months
    list_monthly_sum_revenue = monthly_revenue_df.sum(axis =1)
    for col_idx, value in enumerate(list_monthly_sum_revenue, start = col_index_monthly_data):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting monthly sum for the months
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum_revenue) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")





    # inserting and formatting WEEKLY DATA
    list_columns = [s for s in weekly_sum_df if 'rev_' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    weekly_revenue_df = weekly_sum_df[list_columns]

    #inserting week titles
    for idx, value in enumerate(weekly_revenue_df.index, start = col_index_weekly_data):
        sheet.cell(row = 5, column = idx, value = value)

    for row_idx, row in enumerate(weekly_revenue_df.iterrows(), start = col_index_weekly_data):
        for c_idx, value in enumerate(row[1], start = 8):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = 8, maxRow  = 8 + len(annual_revenue_df.columns) - 1, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")
    
    # formatting annual sum for the weeks
    list_weekly_sum_revenue = weekly_revenue_df.sum(axis =1)
    for col_idx, value in enumerate(list_weekly_sum_revenue, start = col_index_weekly_data):
        sheet.cell(row = 7, column = col_idx, value = value)

    # formatting monthly sum for the weeks
    conditional_formatting_cells(minRow = 7, maxRow  = 7, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")


    # endregion
    # ---------------------------------------------------------------------------------------------------------------------------------------
    # region inserting REVENUE VALUES COMPENSATION


    row_index_comp_data = 7 + 2 + len(annual_revenue_df.columns)

    # operation costs title
    sheet.cell(row = row_index_comp_data, column = 3, value = '2 - Revenue (stock)')

    #formating row operation costs
    formatting_cells(minRow = row_index_comp_data, maxRow = row_index_comp_data, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)

    # inserting and formatting ANNUAL DATA
    list_columns = [s for s in annual_sum_df if 'stock_' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    annual_comp_df = annual_sum_df[list_columns]

    #inserting title of row columns
    for idx, value in enumerate(annual_comp_df.columns, start = row_index_comp_data + 1):
        sheet.cell(row = idx, column = 3, value = value)

    #inserting values of yearly data
    for row_idx, row in enumerate(annual_comp_df.iterrows(), start = 5):
        for c_idx, value in enumerate(row[1], start = row_index_comp_data + 1):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    #formatting titles of row columns
    formatting_cells(minRow = row_index_comp_data + 1, maxRow =  row_index_comp_data + 1 + len(annual_comp_df.columns) - 1, minCol = 1, maxCol = 3, colorCode = "D4EEFC", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style_with_indent)

    #formatting numbers of yearly values
    conditional_formatting_cells(minRow = row_index_comp_data + 1 , maxRow  = row_index_comp_data + 1  + len(annual_comp_df.columns) - 1, minCol = 5, maxCol = 5 + len(annual_sum_df) - 1, 
                                    colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")

    # formatting annual sum for the years
    list_annual_sum_compensation = annual_comp_df.sum(axis =1)
    for col_idx, value in enumerate(list_annual_sum_compensation, start = 5):
        sheet.cell(row = row_index_comp_data, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = row_index_comp_data, maxRow  = row_index_comp_data, minCol = 5, maxCol = 5 + len(list_annual_sum_compensation) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")





    # inserting and formatting MONTHLY DATA
    list_columns = [s for s in monthly_sum_df if 'stock_' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    monthly_comp_df = monthly_sum_df[list_columns]

    #inserting data for monthly values
    for row_idx, row in enumerate(monthly_comp_df.iterrows(), start = col_index_monthly_data):
        for c_idx, value in enumerate(row[1], start = row_index_comp_data + 1 ):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = row_index_comp_data + 1 , maxRow  =  row_index_comp_data + 1 + len(annual_comp_df.columns) - 1, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # inserting annual sum for the months
    list_monthly_sum_compensation = monthly_comp_df.sum(axis =1)
    for col_idx, value in enumerate(list_monthly_sum_compensation, start = col_index_monthly_data):
        sheet.cell(row =  row_index_comp_data, column = col_idx, value = value)

    # formatting monthly sum for the years
    conditional_formatting_cells(minRow =  row_index_comp_data, maxRow  =  row_index_comp_data, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum_compensation) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")




    # inserting and formatting WEEKLY DATA
    list_columns = [s for s in weekly_sum_df if 'stock_' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    weekly_comp_df = weekly_sum_df[list_columns]


    for row_idx, row in enumerate(weekly_comp_df.iterrows(), start = col_index_weekly_data):
        for c_idx, value in enumerate(row[1], start = row_index_comp_data + 1):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = row_index_comp_data + 1 , maxRow  = row_index_comp_data + 1 + len(annual_comp_df.columns) - 1, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # inserting total sum for the weeks
    list_weekly_sum_compensation = weekly_comp_df.sum(axis =1)
    for col_idx, value in enumerate(list_weekly_sum_compensation, start = col_index_weekly_data):
        sheet.cell(row = row_index_comp_data, column = col_idx, value = value)

    # formatting total sum for the weeks
    conditional_formatting_cells(minRow = row_index_comp_data, maxRow  = row_index_comp_data, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(list_weekly_sum_compensation) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")
    

    # endregion
    # ---------------------------------------------------------------------------------------------------------------------------------------
    # region inserting OPERATION COSTS

    row_index_op_costs_data = 7 + 4 + len(annual_revenue_df.columns) + len(annual_comp_df.columns)

    # operation costs title
    sheet.cell(row = row_index_op_costs_data, column = 3, value = '3 - Operational costs')

    #formating row operation costs
    formatting_cells(minRow = row_index_op_costs_data, maxRow = row_index_op_costs_data, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)

    # inserting and formatting ANNUAL DATA
    list_columns = [s for s in annual_sum_df if 'op_costs' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    annual_op_costs_df = annual_sum_df[list_columns]
    annual_op_costs_df *= -1

    #inserting title of row columns
    for idx, value in enumerate(annual_op_costs_df.columns, start = row_index_op_costs_data + 1):
        sheet.cell(row = idx, column = 3, value = value)

    #inserting values of yearly data
    for row_idx, row in enumerate(annual_op_costs_df.iterrows(), start = 5):
        for c_idx, value in enumerate(row[1], start = row_index_op_costs_data + 1):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    #formatting titles of row columns
    formatting_cells(minRow = row_index_op_costs_data + 1, maxRow =  row_index_op_costs_data + 1 + len(annual_op_costs_df.columns) - 1, minCol = 1, maxCol = 3, colorCode = "D4EEFC", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style_with_indent)

    #formatting numbers of yearly values
    conditional_formatting_cells(minRow = row_index_op_costs_data + 1 , maxRow  = row_index_op_costs_data + 1  + len(annual_op_costs_df.columns) - 1, minCol = 5, maxCol = 5 + len(annual_sum_df) - 1, 
                                    colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")

    # formatting annual sum for the years
    list_annual_sum_op_costs = annual_op_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_annual_sum_op_costs, start = 5):
        sheet.cell(row = row_index_op_costs_data, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = row_index_op_costs_data, maxRow  = row_index_op_costs_data, minCol = 5, maxCol = 5 + len(list_annual_sum_op_costs) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")




    # inserting and formatting MONTHLY DATA
    list_columns = [s for s in monthly_sum_df if 'op_costs' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    monthly_op_costs_df = monthly_sum_df[list_columns]
    monthly_op_costs_df *= -1

    #inserting data for monthly values
    for row_idx, row in enumerate(monthly_op_costs_df.iterrows(), start = col_index_monthly_data):
        for c_idx, value in enumerate(row[1], start = row_index_op_costs_data+ 1 ):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = row_index_op_costs_data + 1 , maxRow  =  row_index_op_costs_data + 1 + len(annual_op_costs_df.columns) - 1, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # inserting annual sum for the months
    list_monthly_sum_op_costs = monthly_op_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_monthly_sum_op_costs, start = col_index_monthly_data):
        sheet.cell(row =  row_index_op_costs_data, column = col_idx, value = value)

    # formatting monthly sum for the years
    conditional_formatting_cells(minRow =  row_index_op_costs_data, maxRow  =  row_index_op_costs_data, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum_op_costs) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")




    # inserting and formatting WEEKLY DATA
    list_columns = [s for s in weekly_sum_df if 'op_costs' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    weekly_op_costs_df = weekly_sum_df[list_columns]
    weekly_op_costs_df *= -1


    for row_idx, row in enumerate(weekly_op_costs_df.iterrows(), start = col_index_weekly_data):
        for c_idx, value in enumerate(row[1], start = row_index_op_costs_data + 1):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = row_index_op_costs_data + 1 , maxRow  = row_index_op_costs_data + 1 + len(annual_op_costs_df.columns) - 1, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # inserting total sum for the weeks
    list_weekly_sum_op_costs = weekly_op_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_weekly_sum_op_costs, start = col_index_weekly_data):
        sheet.cell(row = row_index_op_costs_data, column = col_idx, value = value)

    # formatting total sum for the weeks
    conditional_formatting_cells(minRow = row_index_op_costs_data, maxRow  = row_index_op_costs_data, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(list_weekly_sum_op_costs) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")


    # endregion
    # ---------------------------------------------------------------------------------------------------------------------------------------
    # region inserting investment COSTS

    row_index_inv_data = 7 + 6 + len(annual_revenue_df.columns) + len(annual_comp_df.columns) + len(annual_op_costs_df.columns)

    # operation costs title
    sheet.cell(row = row_index_inv_data, column = 3, value = '4 - Investment costs')

    #formating row operation costs
    formatting_cells(minRow = row_index_inv_data, maxRow = row_index_inv_data, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)

    # inserting and formatting ANNUAL DATA
    list_columns = [s for s in annual_sum_df if 'inv_costs' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    annual_inv_costs_df = annual_sum_df[list_columns]
    annual_inv_costs_df *= -1

    #inserting title of row columns
    for idx, value in enumerate(annual_inv_costs_df.columns, start = row_index_inv_data + 1):
        sheet.cell(row = idx, column = 3, value = value)

    #inserting values of yearly data
    for row_idx, row in enumerate(annual_inv_costs_df.iterrows(), start = 5):
        for c_idx, value in enumerate(row[1], start = row_index_inv_data + 1):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    #formatting titles of row columns
    formatting_cells(minRow = row_index_inv_data + 1, maxRow =  row_index_inv_data + 1 + len(annual_inv_costs_df.columns) - 1, minCol = 1, maxCol = 3, colorCode = "D4EEFC", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style_with_indent)

    #formatting numbers of yearly values
    conditional_formatting_cells(minRow = row_index_inv_data + 1 , maxRow  = row_index_inv_data + 1  + len(annual_inv_costs_df.columns) - 1, minCol = 5, maxCol = 5 + len(annual_sum_df) - 1, 
                                    colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")

    # formatting annual sum for the years
    list_annual_sum_inv_costs = annual_inv_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_annual_sum_inv_costs, start = 5):
        sheet.cell(row = row_index_inv_data, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = row_index_inv_data, maxRow  = row_index_inv_data, minCol = 5, maxCol = 5 + len(list_annual_sum_inv_costs) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")





    # inserting and formatting MONTHLY DATA
    list_columns = [s for s in monthly_sum_df if 'inv_costs' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    monthly_inv_costs_df = monthly_sum_df[list_columns]
    monthly_inv_costs_df *= -1

    #inserting data for monthly values
    for row_idx, row in enumerate(monthly_inv_costs_df.iterrows(), start = col_index_monthly_data):
        for c_idx, value in enumerate(row[1], start = row_index_inv_data+ 1 ):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = row_index_inv_data + 1 , maxRow  =  row_index_inv_data + 1 + len(annual_inv_costs_df.columns) - 1, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(monthly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # inserting annual sum for the months
    list_monthly_sum_inv_costs = monthly_inv_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_monthly_sum_inv_costs, start = col_index_monthly_data):
        sheet.cell(row =  row_index_inv_data, column = col_idx, value = value)

    # formatting monthly sum for the years
    conditional_formatting_cells(minRow =  row_index_inv_data, maxRow  =  row_index_inv_data, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum_inv_costs) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")





    # inserting and formatting WEEKLY DATA
    list_columns = [s for s in weekly_sum_df if 'inv_costs' in s]
    list_columns = [s for s in list_columns if 'total' not in s]
    list_columns.sort()
    weekly_inv_costs_df = weekly_sum_df[list_columns]
    weekly_inv_costs_df *= -1


    for row_idx, row in enumerate(weekly_inv_costs_df.iterrows(), start = col_index_weekly_data):
        for c_idx, value in enumerate(row[1], start = row_index_inv_data + 1):
            sheet.cell(row = c_idx, column = row_idx, value = value)

    conditional_formatting_cells(minRow = row_index_inv_data + 1 , maxRow  = row_index_inv_data + 1 + len(annual_inv_costs_df.columns) - 1, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(weekly_sum_df) - 1, 
                                colorCode = "D4EEFC", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # inserting total sum for the weeks
    list_weekly_sum_inv_costs = weekly_inv_costs_df.sum(axis =1)
    for col_idx, value in enumerate(list_weekly_sum_inv_costs, start = col_index_weekly_data):
        sheet.cell(row = row_index_inv_data, column = col_idx, value = value)

    # formatting total sum for the weeks
    conditional_formatting_cells(minRow = row_index_inv_data, maxRow  = row_index_inv_data, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(list_weekly_sum_inv_costs) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # endregion
    # --------------------------------------------------------------------------
    # region other adjustments


    set_column_width(23, 3)
    set_column_width(2, 4)
    set_column_width(2, col_index_monthly_data - 1)
    set_column_width(2, col_index_weekly_data -1)

    for i in range(col_index_weekly_data, col_index_weekly_data + len(weekly_comp_df)):
        set_column_width(13, i)


    # endregion
    # --------------------------------------------------------------------------
    # region FINAL SUM RESULT


    row_index_total = 7 + 8 + len(annual_revenue_df.columns) + len(annual_comp_df.columns) + len(annual_op_costs_df.columns) +len(annual_inv_costs_df.columns)

    # operation costs title
    sheet.cell(row = row_index_total, column = 3, value = '5 - Result')

    #formating row operation costs
    formatting_cells(minRow = row_index_total, maxRow = row_index_total, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)

    annual_result  = np.array(list_annual_sum_revenue) + np.array(list_annual_sum_compensation) + np.array(list_annual_sum_op_costs) + np.array(list_annual_sum_inv_costs)
    monthly_result  = np.array(list_monthly_sum_revenue) + np.array(list_monthly_sum_compensation) + np.array(list_monthly_sum_op_costs) + np.array(list_monthly_sum_inv_costs)
    weekly_result  = np.array(list_weekly_sum_revenue) + np.array(list_weekly_sum_compensation) + np.array(list_weekly_sum_op_costs) + np.array(list_weekly_sum_inv_costs)

    # inserting annual result for the years
    for col_idx, value in enumerate(annual_result, start = 5):
        sheet.cell(row = row_index_total, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = row_index_total, maxRow  = row_index_total, minCol = 5, maxCol = 5 + len(list_annual_sum_revenue) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")


    # inserting annual sum for the months
    for col_idx, value in enumerate(monthly_result, start = col_index_monthly_data):
        sheet.cell(row =  row_index_total, column = col_idx, value = value)

    # formatting monthly sum for the years
    conditional_formatting_cells(minRow =  row_index_total, maxRow  =  row_index_total, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum_revenue) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")


    # inserting total sum for the weeks
    for col_idx, value in enumerate(weekly_result, start = col_index_weekly_data):
        sheet.cell(row = row_index_total, column = col_idx, value = value)

    # formatting total sum for the weeks
    conditional_formatting_cells(minRow = row_index_total, maxRow  = row_index_total, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(list_weekly_sum_revenue) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")
    

    # now inserting the cumulated sum of the cash flows
    row_index_total_cumulated = row_index_total + 1 + 1

    # operation costs title
    sheet.cell(row = row_index_total_cumulated, column = 3, value = '6 - Result accumulated')

    #formating row operation costs
    formatting_cells(minRow = row_index_total_cumulated, maxRow = row_index_total_cumulated, minCol = 1, maxCol = 3, colorCode = "93D3F7", 
                    boldOption = True, sizeOption = 10, fontOption = 'Arial', styleOption = normal_style)
    
    
    # inserting annual result for the years
    for col_idx, value in enumerate(np.cumsum(annual_result), start = 5):
        sheet.cell(row = row_index_total_cumulated, column = col_idx, value = value)

    # formatting annual sum for the years
    conditional_formatting_cells(minRow = row_index_total_cumulated, maxRow  = row_index_total_cumulated, minCol = 5, maxCol = 5 + len(list_annual_sum_revenue) - 1, 
                                    colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                    fontColor = "FF0000")

    # inserting annual sum for the months
    for col_idx, value in enumerate(np.cumsum(monthly_result), start = col_index_monthly_data):
        sheet.cell(row =  row_index_total_cumulated, column = col_idx, value = value)

    # formatting monthly sum for the years
    conditional_formatting_cells(minRow =  row_index_total_cumulated, maxRow  =  row_index_total_cumulated, minCol = col_index_monthly_data, 
                                maxCol = col_index_monthly_data + len(list_monthly_sum_revenue) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")

    # inserting total sum for the weeks
    for col_idx, value in enumerate(np.cumsum(weekly_result), start = col_index_weekly_data):
        sheet.cell(row = row_index_total_cumulated, column = col_idx, value = value)

    # formatting total sum for the weeks
    conditional_formatting_cells(minRow = row_index_total_cumulated, maxRow  = row_index_total_cumulated, minCol = col_index_weekly_data, 
                                maxCol = col_index_weekly_data + len(list_weekly_sum_revenue) - 1, 
                                colorCode = "93D3F7", boldOption = True, fontOption = 'Arial', sizeOption = 10 , styleOption = thousands_separator_format,
                                fontColor = "FF0000")
    
    #endregion

    # Hide gridlines
    sheet.sheet_view.showGridLines = False
    mycell = sheet['D6']
    sheet.freeze_panes = mycell

    # Save the Excel file
    wb.save(path_output + "financial_analysis.xlsx")